#!/usr/bin/env python3
"""
Import edited cashiers from Excel back to database.
Run: python3 import_cashiers.py

Reads: cashiers_export.xlsx
"""

import sqlite3
import json

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip3", "install", "openpyxl"])
    import openpyxl

from openpyxl import load_workbook

DB_PATH = "schedule.db"
INPUT_FILE = "cashiers_export.xlsx"

def parse_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).lower().strip() in ('yes', 'true', '1', 'y')

def main():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Load Excel
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    days = ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat']
    shifts = ['1st', '2nd', '3rd']
    
    updated = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            emp_id = row[0]
            if not emp_id:
                continue
            
            first_name = row[1] or ''
            last_name = row[2] or ''
            phone = row[3] or ''
            email = row[4] or ''
            hire_date = row[5] or None
            is_full_time = parse_bool(row[6])
            max_weekly_hours = row[7] or 40
            notes = row[8] or ''
            no_overtime = parse_bool(row[9])
            no_plaza = parse_bool(row[10])
            is_active = parse_bool(row[11])
            
            # Parse availability grid from columns 12+
            grid = {}
            col_idx = 12
            for day in days:
                grid[day] = {}
                for shift in shifts:
                    if col_idx < len(row):
                        grid[day][shift] = parse_bool(row[col_idx])
                    else:
                        grid[day][shift] = True
                    col_idx += 1
            
            availability_grid = json.dumps(grid)
            
            # Update database
            cursor.execute("""
                UPDATE employee SET
                    first_name = ?,
                    last_name = ?,
                    phone = ?,
                    email = ?,
                    hire_date = ?,
                    is_full_time = ?,
                    max_weekly_hours = ?,
                    notes = ?,
                    no_overtime = ?,
                    no_plaza = ?,
                    is_active = ?,
                    availability_grid = ?
                WHERE id = ?
            """, (
                first_name, last_name, phone, email, hire_date,
                1 if is_full_time else 0,
                max_weekly_hours, notes,
                1 if no_overtime else 0,
                1 if no_plaza else 0,
                1 if is_active else 0,
                availability_grid,
                emp_id
            ))
            
            updated += 1
            print(f"Updated: {first_name} {last_name} (ID: {emp_id})")
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")
    
    conn.commit()
    conn.close()
    
    print(f"\n✅ Updated {updated} employees")
    if errors:
        print(f"\n⚠️  Errors ({len(errors)}):")
        for err in errors[:10]:
            print(f"  - {err}")

if __name__ == "__main__":
    main()
