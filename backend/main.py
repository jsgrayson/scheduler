from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from sqlmodel import SQLModel, Session, select, create_engine, delete
from typing import List, Optional
from datetime import datetime, timedelta, time
from database import create_db_and_tables, get_session
from models import Employee, Role, Shift, Availability, EmployeeRole, EmployeeBase, RotationState
from pydantic import BaseModel

app = FastAPI()

# CORS Middleware (allow all for local dev)
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

KNOWN_LOCATIONS = [
    "LOT 1", "LOT 2", "LOT 3", "LOT 4", 
    "PLAZA", "CONRAC", "OFFICE", "MAINTENANCE", 
    "SUPERVISORS", "CUSTOMER LOTS", "CASHIER"
]

LOCATION_MAPPINGS = {
    "SUP3": "Supervisors",
    # "SUP": "Supervisors", # Too ambiguous, matches "Jim E (Sup)"
    "FLOAT": "Office", 
    "OFF: MGD": "Office",
    "RECPTAR": "Office",
    "ADIAU": "Office",
    "ADMIN": "Office",
    "AFM": "Office",
    "AD": "Office",
    "SUP-MGR": "Maintenance",
    # "SUP": "Maintenance", # Too ambiguous
    "C-LOT": "Customer Lots",
    "CLOT": "Customer Lots",
    "E-LOT": "Customer Lots",
    "ELOT": "Customer Lots",
    "CUSTOMER LOT": "Customer Lots",
}



@app.on_event("startup")
def on_startup():
    # Database Backup
    import shutil
    import os
    try:
        db_file = "schedule.db"
        if os.path.exists(db_file):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"backups/schedule_backup_{timestamp}.db"
            os.makedirs("backups", exist_ok=True)
            shutil.copy2(db_file, backup_name)
            print(f"Database backed up to {backup_name}")
            
            # Keep only last 10 backups
            backups = sorted([f for f in os.listdir("backups") if f.endswith(".db")])
            if len(backups) > 10:
                for b in backups[:-10]:
                    os.remove(os.path.join("backups", b))
                    print(f"Removed old backup {b}")
    except Exception as e:
        print(f"Backup failed: {e}")

    create_db_and_tables()

from pydantic import BaseModel, ConfigDict

# --- Employees ---
class EmployeeRead(EmployeeBase):
    id: int
    roles: List[Role] = []

@app.get("/employees/", response_model=List[EmployeeRead])
def read_employees(session: Session = Depends(get_session)):
    employees = session.exec(select(Employee)).all()
    return employees

@app.post("/employees/", response_model=Employee)
def create_employee(employee: Employee, session: Session = Depends(get_session)):
    session.add(employee)
    session.commit()
    session.refresh(employee)
    return employee

# --- Roles ---
class EmployeeUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    is_full_time: Optional[bool] = None
    default_role_id: Optional[int] = None
    willing_to_work_vacation_week: Optional[bool] = None
    max_weekly_hours: Optional[float] = None
    hire_date: Optional[datetime] = None
    role_ids: Optional[List[int]] = None # New field for multi-role

@app.put("/employees/{employee_id}", response_model=Employee)
def update_employee(employee_id: int, employee_data: EmployeeUpdate, session: Session = Depends(get_session)):
    employee = session.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    hero_data = employee_data.model_dump(exclude_unset=True)
    
    # Handle role_ids separately
    if "role_ids" in hero_data:
        role_ids = hero_data.pop("role_ids")
        # Clear existing roles
        session.exec(delete(EmployeeRole).where(EmployeeRole.employee_id == employee_id))
        # Add new roles
        for rid in role_ids:
            session.add(EmployeeRole(employee_id=employee_id, role_id=rid))
            
    for key, value in hero_data.items():
        setattr(employee, key, value)
        
    session.add(employee)
    session.commit()
    session.refresh(employee)
    return employee

@app.get("/roles/", response_model=List[Role])
def read_roles(session: Session = Depends(get_session)):
    roles = session.exec(select(Role)).all()
    return roles

# --- Shifts ---
@app.get("/shifts/", response_model=List[Shift])
def read_shifts(
    start_date: datetime,
    end_date: datetime,
    session: Session = Depends(get_session)
):
    # Get shifts that overlap with the date range (not strictly within)
    # A shift overlaps if: shift_start < range_end AND shift_end > range_start
    statement = select(Shift).where(Shift.start_time < end_date).where(Shift.end_time > start_date)
    shifts = session.exec(statement).all()
    return shifts

class ShiftCreate(BaseModel):
    employee_id: Optional[int] = None
    role_id: int
    start_time: datetime
    end_time: datetime
    notes: Optional[str] = None
    location: Optional[str] = None
    booth_number: Optional[str] = None
    is_vacation: bool = False
    repeat: Optional[str] = None # "daily", "weekly", "mon-fri"
    create_open_shift: bool = False # If vacation, create covering open shift

class ShiftRead(BaseModel):
    id: int
    employee_id: Optional[int]
    role_id: int
    start_time: datetime
    end_time: datetime
    notes: Optional[str] = None
    location: Optional[str] = None
    booth_number: Optional[str] = None
    is_vacation: bool
    is_locked: bool = False
    parent_id: Optional[int] = None

@app.post("/shifts/", response_model=List[ShiftRead])
def create_shift(shift_data: ShiftCreate, session: Session = Depends(get_session)):
    # Validate end time is after start time
    if shift_data.end_time <= shift_data.start_time:
        raise HTTPException(status_code=400, detail="End time must be after start time")
    
    # 1. Base Shift Data
    shifts_to_create = []
    
    # Calculate duration
    duration = shift_data.end_time - shift_data.start_time
    
    # Determine start dates based on recurrence
    start_dates = [shift_data.start_time]
    
    if shift_data.repeat:
        # Generate for next 4 weeks (28 days)
        current = shift_data.start_time
        for _ in range(28): # Check next 28 days
            current += timedelta(days=1)
            
            should_add = False
            if shift_data.repeat == "daily":
                should_add = True
            elif shift_data.repeat == "weekly":
                if current.weekday() == shift_data.start_time.weekday():
                    should_add = True
            elif shift_data.repeat == "mon-fri":
                if current.weekday() < 5:
                    should_add = True
            
            if should_add:
                start_dates.append(current)
                
    # Create Shift Objects
    parent_id = None # Will be set to first shift's ID if repeating
    
    created_shifts = []
    
    for i, start_dt in enumerate(start_dates):
        end_dt = start_dt + duration
        
        # Conflict Check (Skip for Vacation? Maybe warn but allow? Let's check for now)
        # If vacation, we might want to allow it even if there's a work shift (and maybe delete work shift?)
        # For MVP, simple check.
        
        # Create Shift
        shift = Shift(
            employee_id=shift_data.employee_id,
            role_id=shift_data.role_id,
            start_time=start_dt,
            end_time=end_dt,
            notes=shift_data.notes,
            location=shift_data.location,
            booth_number=shift_data.booth_number,
            is_vacation=shift_data.is_vacation,
            parent_id=parent_id
        )
        session.add(shift)
        session.commit()
        session.refresh(shift)
        
        if i == 0 and len(start_dates) > 1:
            parent_id = shift.id
            shift.parent_id = shift.id # Self-reference for parent
            session.add(shift)
            session.commit()
            
        created_shifts.append(shift)
        
        # Handle Vacation Cover
        if shift_data.is_vacation and shift_data.create_open_shift:
            # Create Open Shift
            open_shift = Shift(
                employee_id=None, # Open
                role_id=shift_data.role_id,
                start_time=start_dt,
                end_time=end_dt,
                notes=f"Cover for {shift_data.notes or 'Vacation'}"
            )
            session.add(open_shift)
            session.commit()
            created_shifts.append(open_shift)
            
            created_shifts.append(open_shift)
            
    return created_shifts

@app.get("/shifts/agenda/{employee_id}", response_model=List[Shift])
def get_agenda(employee_id: int, session: Session = Depends(get_session)):
    # Get future shifts for employee
    now = datetime.now()
    statement = select(Shift).where(
        Shift.employee_id == employee_id,
    ).order_by(Shift.start_time)
    shifts = session.exec(statement).all()
    return shifts

class ShiftUpdate(BaseModel):
    employee_id: Optional[int] = None
    role_id: Optional[int] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    notes: Optional[str] = None
    location: Optional[str] = None
    booth_number: Optional[str] = None
    is_vacation: Optional[bool] = None
    is_locked: Optional[bool] = None
    force_save: Optional[bool] = False  # Skip conflict detection

@app.get("/shifts/{shift_id}", response_model=ShiftRead)
def get_shift(shift_id: int, session: Session = Depends(get_session)):
    shift = session.get(Shift, shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    return shift

@app.put("/shifts/{shift_id}", response_model=Shift)
def update_shift(shift_id: int, shift_data: ShiftUpdate, session: Session = Depends(get_session)):
    shift = session.get(Shift, shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    # Update only provided fields
    update_data = shift_data.model_dump(exclude_unset=True)
    force_save = update_data.pop('force_save', False)
    
    # Check if shift is locked (unless we're just unlocking it)
    is_unlocking = update_data.get('is_locked') == False
    if shift.is_locked and not is_unlocking and not force_save:
        raise HTTPException(status_code=403, detail="Shift is locked. Use Force Save to override.")
    
    # Conflict detection (excluding self, and only if employee assigned) - skip if force_save
    if not force_save and 'employee_id' in update_data and update_data['employee_id']:
        start_time = update_data.get('start_time', shift.start_time)
        end_time = update_data.get('end_time', shift.end_time)
        
        # Validate end time is after start time
        if end_time <= start_time:
            raise HTTPException(status_code=400, detail="End time must be after start time")
        
        statement = select(Shift).where(
            Shift.employee_id == update_data['employee_id'],
            Shift.start_time < end_time,
            Shift.end_time > start_time,
            Shift.id != shift_id
        )
        conflicts = session.exec(statement).all()
        if conflicts:
            raise HTTPException(status_code=400, detail="Shift overlaps with an existing shift.")
    
    # Apply updates
    for key, value in update_data.items():
        setattr(shift, key, value)
    
    session.add(shift)
    session.commit()
    session.refresh(shift)
    return shift

@app.delete("/shifts/{shift_id}")
def delete_shift(shift_id: int, force: bool = False, session: Session = Depends(get_session)):
    shift = session.get(Shift, shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    if shift.is_locked and not force:
        raise HTTPException(status_code=403, detail="Shift is locked. Cannot delete.")
    session.delete(shift)
    session.commit()
    return {"ok": True}

# --- Bulk Update Shifts ---
class BulkShiftUpdate(BaseModel):
    shift_ids: List[int]
    role_id: Optional[int] = None
    location: Optional[str] = None
    booth_number: Optional[str] = None
    is_locked: Optional[bool] = None

@app.post("/shifts/bulk-update/")
def bulk_update_shifts(data: BulkShiftUpdate, session: Session = Depends(get_session)):
    if not data.shift_ids:
        raise HTTPException(status_code=400, detail="No shift IDs provided")
    
    if data.role_id is None and data.location is None and data.booth_number is None and data.is_locked is None:
        raise HTTPException(status_code=400, detail="No updates specified")
    
    updated_count = 0
    for shift_id in data.shift_ids:
        shift = session.get(Shift, shift_id)
        if shift:
            if data.role_id is not None:
                shift.role_id = data.role_id
            if data.location is not None:
                shift.location = data.location
            if data.booth_number is not None:
                shift.booth_number = data.booth_number
            if data.is_locked is not None:
                shift.is_locked = data.is_locked
            session.add(shift)
            updated_count += 1
    
    session.commit()
    return {"ok": True, "updated_count": updated_count}

# --- Bulk Delete Shifts ---
class BulkShiftDelete(BaseModel):
    shift_ids: List[int]

@app.post("/shifts/bulk-delete/")
def bulk_delete_shifts(data: BulkShiftDelete, session: Session = Depends(get_session)):
    if not data.shift_ids:
        raise HTTPException(status_code=400, detail="No shift IDs provided")
    
    deleted_count = 0
    for shift_id in data.shift_ids:
        shift = session.get(Shift, shift_id)
        if shift:
            session.delete(shift)
            deleted_count += 1
    
    session.commit()
    return {"ok": True, "deleted_count": deleted_count}

# --- Project Locked Shifts to Future Weeks ---
class ProjectLockedRequest(BaseModel):
    base_week_start: datetime  # Saturday of the base week
    num_weeks: int = 4  # Number of future weeks to project to

@app.post("/shifts/project-locked/")
def project_locked_shifts(data: ProjectLockedRequest, session: Session = Depends(get_session)):
    from datetime import timedelta
    
    # Get all locked shifts in the base week
    base_week_end = data.base_week_start + timedelta(days=7)
    locked_shifts = session.exec(
        select(Shift).where(
            Shift.is_locked == True,
            Shift.start_time >= data.base_week_start,
            Shift.start_time < base_week_end
        )
    ).all()
    
    if not locked_shifts:
        raise HTTPException(status_code=400, detail="No locked shifts found in the base week")
    
    created_count = 0
    updated_count = 0
    
    for week_offset in range(1, data.num_weeks + 1):
        days_offset = timedelta(days=7 * week_offset)
        
        for base_shift in locked_shifts:
            new_start = base_shift.start_time + days_offset
            new_end = base_shift.end_time + days_offset
            
            # Target day range
            target_day_start = new_start.replace(hour=0, minute=0, second=0, microsecond=0)
            target_day_end = target_day_start + timedelta(days=1)
            
            # Check for existing shifts for this employee on this day
            existing_shifts = session.exec(
                select(Shift).where(
                    Shift.employee_id == base_shift.employee_id,
                    Shift.start_time >= target_day_start,
                    Shift.start_time < target_day_end
                )
            ).all()
            
            # Logic:
            # 1. If any existing shift is LOCKED, skip this projection (conflict with another master)
            # 2. If existing shifts are UNLOCKED, delete them (assume we are overwriting with new master)
            
            has_locked_conflict = any(s.is_locked for s in existing_shifts)
            if has_locked_conflict:
                continue # Skip, respect future lock
            
            # Delete existing unlocked shifts on this day to prevent recurrence duplicates
            for s in existing_shifts:
                session.delete(s)
                updated_count += 1 # Count deletions as updates
            
            # Create new projected shift
            new_shift = Shift(
                employee_id=base_shift.employee_id,
                role_id=base_shift.role_id,
                start_time=new_start,
                end_time=new_end,
                notes=base_shift.notes,
                location=base_shift.location,
                booth_number=base_shift.booth_number,
                is_locked=False,  # Projected shifts are not locked
                is_vacation=base_shift.is_vacation
            )
            session.add(new_shift)
            created_count += 1
    
    session.commit()
    return {"ok": True, "created_count": created_count, "deleted_old_count": updated_count, "weeks_projected": data.num_weeks}

# --- Shift Templates (Master Schedule) ---
from models import ShiftTemplate

@app.get("/templates/", response_model=List[ShiftTemplate])
def read_templates(session: Session = Depends(get_session)):
    return session.exec(select(ShiftTemplate)).all()

class ShiftTemplateRequest(BaseModel):
    employee_id: int
    role_id: int
    day_of_week: int
    start_time: str
    end_time: str
    location: Optional[str] = None
    booth_number: Optional[str] = None
    sync_to_locked: bool = False

@app.post("/templates/", response_model=ShiftTemplate)
def create_template(data: ShiftTemplateRequest, session: Session = Depends(get_session)):
    template = ShiftTemplate(
        employee_id=data.employee_id,
        role_id=data.role_id,
        day_of_week=data.day_of_week,
        start_time=data.start_time,
        end_time=data.end_time,
        location=data.location,
        booth_number=data.booth_number
    )
    session.add(template)
    session.commit()
    session.refresh(template)
    
    if data.sync_to_locked:
        from datetime import date
        today = datetime.now().date()
        
        # Calculate first occurrence of the target day
        days_ahead = data.day_of_week - today.weekday()
        if days_ahead < 0: 
            days_ahead += 7
        
        start_date = today + timedelta(days=days_ahead)
        
        # Sync for 8 weeks
        for i in range(8):
            target_date = start_date + timedelta(weeks=i)
            
            # Find ANY shift for this employee on this day
            day_start = datetime.combine(target_date, time(0,0))
            day_end = datetime.combine(target_date, time(23,59,59))
            
            existing_shift = session.exec(select(Shift).where(
                Shift.employee_id == data.employee_id,
                Shift.start_time >= day_start,
                Shift.start_time <= day_end
            )).first()
            
            # Prepare new times
            th, tm = map(int, data.start_time.split(':'))
            eh, em = map(int, data.end_time.split(':'))
            
            new_start_dt = datetime.combine(target_date, time(th, tm))
            
            if data.start_time > data.end_time:
                end_date = target_date + timedelta(days=1)
            else:
                end_date = target_date
            
            new_end_dt = datetime.combine(end_date, time(eh, em))
            
            if existing_shift:
                # Update and Lock
                existing_shift.start_time = new_start_dt
                existing_shift.end_time = new_end_dt
                existing_shift.role_id = data.role_id
                existing_shift.location = data.location
                existing_shift.booth_number = data.booth_number
                existing_shift.is_locked = True
                session.add(existing_shift)
            else:
                # Create New
                new_shift = Shift(
                    employee_id=data.employee_id,
                    role_id=data.role_id,
                    start_time=new_start_dt,
                    end_time=new_end_dt,
                    location=data.location,
                    booth_number=data.booth_number,
                    is_locked=True,
                    is_repeating=False
                )
                session.add(new_shift)
                
        session.commit()

    return template

@app.delete("/templates/{template_id}")
def delete_template(template_id: int, session: Session = Depends(get_session)):
    template = session.get(ShiftTemplate, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    session.delete(template)
    session.commit()
    return {"ok": True}

class ApplyScheduleRequest(BaseModel):
    start_date: datetime
    num_weeks: int = 4

@app.post("/shifts/apply-schedule/")
def apply_schedule(data: ApplyScheduleRequest, session: Session = Depends(get_session)):
    # Applies templates to generate shifts
    templates = session.exec(select(ShiftTemplate)).all()
    if not templates:
        raise HTTPException(status_code=400, detail="No templates found")
        
    created_count = 0
    from datetime import timedelta
    
    # Loop through weeks
    for week_i in range(data.num_weeks):
        # Calculate Monday of this week
        # Assumes start_date is the start of the period
        week_start = data.start_date + timedelta(weeks=week_i)
        
        for tmpl in templates:
            # Calculate target date: Week Start + Day of Week
            # Note: template.day_of_week 0=Monday. 
            # If week_start is Saturday, we need to adjust.
            # Let's assume week_start provided is the "Start of Week" (e.g. Saturday)
            # And our Week definition is Sat=0? Or Mon=0?
            # Model says 0=Monday.
            # Week starts on Saturday (5). Monday is (0).
            # So if start_date is Saturday:
            # Sat(5), Sun(6), Mon(0), Tue(1)...
            
            # Simple approach: Find the date that matches the day_of_week within the 7 days following week_start containing
            
            for day_offset in range(7):
                current_day = week_start + timedelta(days=day_offset)
                if current_day.weekday() == tmpl.day_of_week:
                    # Match! Create shift
                    
                    # Parse HH:MM
                    th, tm = map(int, tmpl.start_time.split(':'))
                    eh, em = map(int, tmpl.end_time.split(':'))
                    
                    s_dt = current_day.replace(hour=th, minute=tm, second=0, microsecond=0)
                    e_dt = current_day.replace(hour=eh, minute=em, second=0, microsecond=0)
                    
                    # Handle overnight shifts (end < start)
                    if e_dt <= s_dt:
                        e_dt += timedelta(days=1)
                        
                    # Check for existing locked shift? (Respect Manual Locks)
                    existing = session.exec(select(Shift).where(
                        Shift.employee_id == tmpl.employee_id,
                        Shift.start_time == s_dt,
                        Shift.is_locked == True
                    )).first()
                    
                    if not existing:
                        # Create locked shift from template
                        shift = Shift(
                            employee_id=tmpl.employee_id,
                            role_id=tmpl.role_id,
                            start_time=s_dt,
                            end_time=e_dt,
                            location=tmpl.location,
                            booth_number=tmpl.booth_number,
                            is_repeating=False,
                            is_locked=True # Template shifts are locked by default
                        )
                        session.add(shift)
                        created_count += 1
                    break
    
    session.commit()
    return {"ok": True, "created_count": created_count}

@app.post("/templates/import-from-locked/")
def import_templates_from_locked(week_start: datetime, session: Session = Depends(get_session)):
    # Find locked shifts in this week
    week_end = week_start + timedelta(days=7)
    locked_shifts = session.exec(select(Shift).where(
        Shift.is_locked == True,
        Shift.start_time >= week_start,
        Shift.start_time < week_end
    )).all()
    
    if not locked_shifts:
        raise HTTPException(status_code=400, detail="No locked shifts found in this week")
        
    created_count = 0
    skipped_count = 0
    
    for s in locked_shifts:
        # Check if template already exists (same emp, day, time)
        day_of_week = s.start_time.weekday()
        s_time = s.start_time.strftime("%H:%M")
        e_time = s.end_time.strftime("%H:%M")
        
        existing = session.exec(select(ShiftTemplate).where(
            ShiftTemplate.employee_id == s.employee_id,
            ShiftTemplate.day_of_week == day_of_week,
            ShiftTemplate.start_time == s_time
        )).first()
        
        if not existing:
            tmpl = ShiftTemplate(
                employee_id=s.employee_id,
                role_id=s.role_id,
                day_of_week=day_of_week,
                start_time=s_time,
                end_time=e_time,
                location=s.location,
                booth_number=s.booth_number
            )
            session.add(tmpl)
            created_count += 1
        else:
            skipped_count += 1
            
    session.commit()
    return {"ok": True, "created": created_count, "skipped": skipped_count}

# --- Availability ---
from models import Availability

@app.get("/availability/", response_model=List[Availability])
def read_availability(employee_id: Optional[int] = None, session: Session = Depends(get_session)):
    if employee_id:
        statement = select(Availability).where(Availability.employee_id == employee_id)
    else:
        statement = select(Availability)
    return session.exec(statement).all()

@app.post("/availability/", response_model=Availability)
def create_availability(avail: Availability, session: Session = Depends(get_session)):
    session.add(avail)
    session.commit()
    session.refresh(avail)
    return avail

# --- Auto-Scheduler ---
@app.post("/shifts/autofill/", response_model=List[Shift])
def autofill_shifts(session: Session = Depends(get_session)):
    # 1. Get all open shifts
    open_shifts = session.exec(select(Shift).where(Shift.employee_id == None)).all()
    
    filled_shifts = []
    
    for shift in open_shifts:
        # Get potential candidates with matching role
        candidates = session.exec(select(Employee).where(Employee.default_role_id == shift.role_id)).all()
        
        best_candidate = None
        
        for employee in candidates:
            # --- Check 1: Conflicts ---
            conflict = session.exec(select(Shift).where(
                Shift.employee_id == employee.id,
                Shift.start_time < shift.end_time,
                Shift.end_time > shift.start_time
            )).first()
            if conflict:
                continue

            # --- Check 2: Availability ---
            # Get availability for this day of week (0=Mon, 6=Sun)
            day_of_week = shift.start_time.weekday()
            availabilities = session.exec(select(Availability).where(
                Availability.employee_id == employee.id,
                Availability.day_of_week == day_of_week
            )).all()
            
            is_available = True
            
            # If there are "Available" (True) records, shift MUST fit inside one
            positive_avail = [a for a in availabilities if a.is_available]
            if positive_avail:
                fits_in_slot = False
                shift_start_str = shift.start_time.strftime("%H:%M")
                shift_end_str = shift.end_time.strftime("%H:%M")
                
                for slot in positive_avail:
                    if slot.start_time <= shift_start_str and slot.end_time >= shift_end_str:
                        fits_in_slot = True
                        break
                if not fits_in_slot:
                    is_available = False
            
            # If there are "Unavailable" (False) records, shift MUST NOT overlap
            negative_avail = [a for a in availabilities if not a.is_available]
            for slot in negative_avail:
                shift_start_str = shift.start_time.strftime("%H:%M")
                shift_end_str = shift.end_time.strftime("%H:%M")
                
                # Check overlap: start < slot_end AND end > slot_start
                if shift_start_str < slot.end_time and shift_end_str > slot.start_time:
                    is_available = False
                    break
            
            if not is_available:
                continue

            # --- Check 3: Weekly Hours Cap ---
            if employee.max_weekly_hours:
                shift_start = shift.start_time
                start_of_week = shift_start - timedelta(days=shift_start.weekday())
                start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_week = start_of_week + timedelta(days=7)
                
                weekly_shifts = session.exec(select(Shift).where(
                    Shift.employee_id == employee.id,
                    Shift.start_time >= start_of_week,
                    Shift.end_time < end_of_week
                )).all()
                
                current_hours = sum((s.end_time - s.start_time).total_seconds() / 3600 for s in weekly_shifts)
                new_shift_hours = (shift.end_time - shift.start_time).total_seconds() / 3600
                
                if current_hours + new_shift_hours > employee.max_weekly_hours:
                    continue

            # If we passed all checks, assign this employee
            # (Simple logic: take the first valid one. Could be optimized to balance hours)
            best_candidate = employee
            break
        
        if best_candidate:
            shift.employee_id = best_candidate.id
            session.add(shift)
            filled_shifts.append(shift)
    
    session.commit()
    for s in filled_shifts:
        session.refresh(s)
        
    return filled_shifts

# --- Excel Import ---
from io import BytesIO
import openpyxl
from dateutil import parser

@app.post("/import/excel/")
async def import_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    sheet = wb.active
    
    imported_count = 0
    errors = []
    
    # Assuming headers in row 1: Employee, Role, Date, Start Time, End Time, Notes
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]: continue # Skip empty rows
        
        emp_name, role_name, date_str, start_str, end_str, notes = row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None
        
        # 1. Find Employee
        # Split name "First Last"
        parts = str(emp_name).split(" ", 1)
        first_name = parts[0]
        last_name = parts[1] if len(parts) > 1 else ""
        
        print(f"DEBUG: Looking for Employee '{first_name}' '{last_name}'")
        employee = session.exec(select(Employee).where(Employee.first_name == first_name, Employee.last_name == last_name)).first()
        print(f"DEBUG: Found: {employee}")
        
        if not employee:
            errors.append(f"Row {i}: Employee '{emp_name}' not found.")
            continue
            
        # 2. Find Role
        role = session.exec(select(Role).where(Role.name == role_name)).first()
        if not role:
            errors.append(f"Row {i}: Role '{role_name}' not found.")
            continue
            
        # 3. Parse Times
        try:
            # Handle various date/time formats from Excel
            # If date_str is already a datetime object (openpyxl does this), use it
            base_date = date_str if isinstance(date_str, datetime) else parser.parse(str(date_str))
            
            # Combine date with time
            # start_str might be a time object or string
            if isinstance(start_str, datetime): # Sometimes Excel returns datetime for time cells
                start_time = start_str
            elif hasattr(start_str, 'hour'): # time object
                start_time = base_date.replace(hour=start_str.hour, minute=start_str.minute)
            else:
                t = parser.parse(str(start_str)).time()
                start_time = base_date.replace(hour=t.hour, minute=t.minute)
                
            if isinstance(end_str, datetime):
                end_time = end_str
            elif hasattr(end_str, 'hour'):
                end_time = base_date.replace(hour=end_str.hour, minute=end_str.minute)
            else:
                t = parser.parse(str(end_str)).time()
                end_time = base_date.replace(hour=t.hour, minute=t.minute)
                
            # Handle overnight shifts (end time < start time)
            if end_time < start_time:
                end_time += timedelta(days=1)
                
        except Exception as e:
            errors.append(f"Row {i}: Invalid date/time format. {e}")
            continue
            
        # Create Shift
        shift = Shift(
            employee_id=employee.id,
            role_id=role.id,
            start_time=start_time,
            end_time=end_time,
            notes=notes
        )
        session.add(shift)
        imported_count += 1
        
    # Commit whatever was added (partial success)
    session.commit()
    
    # Return success with list of errors if any
    return {"message": "Import process completed", "imported_count": imported_count, "errors": errors}

# --- Call Sheet / Validation ---
from pydantic import BaseModel

class ValidationRequest(BaseModel):
    shifts: List[Shift] # List of proposed shifts (some might not have IDs yet)

@app.post("/shifts/validate/")
def validate_shifts(request: ValidationRequest, session: Session = Depends(get_session)):
    report = {
        "valid": True,
        "conflicts": [],
        "overtime_warnings": []
    }
    
    # Group proposed shifts by employee for OT calc
    emp_hours = {} # {emp_id: total_hours}
    
    # 1. Pre-load existing hours for the relevant week(s)
    # For simplicity, we'll just calculate based on the proposed shifts + existing shifts in DB
    # This is a complex check, let's do a per-shift check
    
    for idx, shift in enumerate(request.shifts):
        if not shift.employee_id: continue
        
        # Ensure datetimes
        if isinstance(shift.start_time, str):
            shift.start_time = parser.parse(shift.start_time)
        if isinstance(shift.end_time, str):
            shift.end_time = parser.parse(shift.end_time)
            
        # --- Conflict Check ---
        # Check against DB
        db_conflicts = session.exec(select(Shift).where(
            Shift.employee_id == shift.employee_id,
            Shift.start_time < shift.end_time,
            Shift.end_time > shift.start_time,
            Shift.id != shift.id # Exclude self if updating
        )).all()
        
        if db_conflicts:
            report["valid"] = False
            report["conflicts"].append(f"Shift {idx+1} overlaps with existing shift ID {[s.id for s in db_conflicts]}")
            
        # Check against other proposed shifts in this batch
        for other_idx, other_shift in enumerate(request.shifts):
            if idx == other_idx: continue
            if other_shift.employee_id != shift.employee_id: continue
            
            # Ensure datetimes for other shift too (might be redundant but safe)
            other_start = other_shift.start_time
            if isinstance(other_start, str): other_start = parser.parse(other_start)
            other_end = other_shift.end_time
            if isinstance(other_end, str): other_end = parser.parse(other_end)
            
            if (shift.start_time < other_end and shift.end_time > other_start):
                 report["valid"] = False
                 report["conflicts"].append(f"Shift {idx+1} overlaps with proposed shift {other_idx+1}")

        # --- Overtime Check ---
        # Calculate hours for this shift
        duration = (shift.end_time - shift.start_time).total_seconds() / 3600
        
        # Add to accumulator
        if shift.employee_id not in emp_hours:
            # Fetch existing hours for this week from DB
            employee = session.get(Employee, shift.employee_id)
            if not employee: continue
            
            shift_start = shift.start_time
            start_of_week = shift_start - timedelta(days=shift_start.weekday())
            start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_week = start_of_week + timedelta(days=7)
            
            weekly_shifts = session.exec(select(Shift).where(
                Shift.employee_id == shift.employee_id,
                Shift.start_time >= start_of_week,
                Shift.end_time < end_of_week
            )).all()
            
            existing_hours = sum((s.end_time - s.start_time).total_seconds() / 3600 for s in weekly_shifts)
            emp_hours[shift.employee_id] = {
                "total": existing_hours,
                "limit": employee.max_weekly_hours or 40,
                "name": f"{employee.first_name} {employee.last_name}"
            }
            
        emp_hours[shift.employee_id]["total"] += duration

    # Generate OT warnings
    for emp_id, data in emp_hours.items():
        if data["total"] > data["limit"]:
            report["overtime_warnings"].append(
                f"Employee {data['name']} is projected to work {data['total']:.1f} hours (Limit: {data['limit']})"
            )
            
    return report

# --- Smart Recommendations ---
@app.get("/recommendations/")
def get_recommendations(
    start_time: datetime,
    end_time: datetime,
    role_id: Optional[int] = None,
    session: Session = Depends(get_session)
):
    """
    Returns a ranked list of employees for a specific time slot.
    Ranking criteria:
    1. Available (Must be true)
    2. Role Match (Must be true if role_id provided)
    3. No Conflict (Must be true)
    4. Weekly Hours (Prefer < Max)
    5. Daily Hours (Prefer < 8)
    """
    
    # Get candidates
    if role_id:
        candidates = session.exec(select(Employee).where(Employee.default_role_id == role_id)).all()
    else:
        candidates = session.exec(select(Employee)).all()
    # 2. Filter and Score Employees
    recommendations = []
    
    # Calculate start/end of the week for the proposed shift
    # Assuming week starts on Monday
    shift_date = start_time.date()
    start_of_week = shift_date - timedelta(days=shift_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    week_start_dt = datetime.combine(start_of_week, datetime.min.time())
    week_end_dt = datetime.combine(end_of_week, datetime.max.time())

    for emp in candidates:
        score = 100
        reasons = []
        is_valid = True
        
        # Check for Vacation in the current week
        # We need to query if this employee has ANY vacation shift this week
        vacation_shifts = session.exec(select(Shift).where(
            Shift.employee_id == emp.id,
            Shift.is_vacation == True,
            Shift.start_time >= week_start_dt,
            Shift.end_time <= week_end_dt
        )).all()
        
        has_vacation_this_week = len(vacation_shifts) > 0
        
        if has_vacation_this_week:
            if emp.is_full_time:
                # FT: No calls if on vacation this week
                is_valid = False
                reasons.append("Full-time employee on vacation this week")
            else:
                # PT: Only if willing
                if not emp.willing_to_work_vacation_week:
                    is_valid = False
                    reasons.append("Part-time employee on vacation this week and not willing to work")
                # If willing, ensure no direct conflict (handled below) and strictly non-vacation day?
                # The direct conflict check below handles the specific time slot.
                # User said: "only on their non vacation shift days".
                # If they have a vacation shift TODAY, they are busy.
                # The conflict check below will catch if the vacation shift overlaps with the proposed time.
                # But if the vacation shift is all day (e.g. 9-5) and we ask for 6-10?
                # If is_vacation is True, usually implies unavailable for work that day?
                # Current implementation: Vacation is a shift with start/end.
                # So conflict check works.
                if is_valid: # Only add reason if still valid
                    reasons.append("Willing to work during vacation week")

        if not is_valid:
            continue # Skip invalid candidates
            
        # 1. Conflict Check
        conflict = session.exec(select(Shift).where(
            Shift.employee_id == emp.id,
            Shift.start_time < end_time,
            Shift.end_time > start_time
        )).first()
        if conflict:
            is_valid = False
            reasons.append("Has conflicting shift")
            
        # 2. Availability Check
        day_of_week = start_time.weekday()
        availabilities = session.exec(select(Availability).where(
            Availability.employee_id == emp.id,
            Availability.day_of_week == day_of_week
        )).all()
        
        # Logic: If ANY positive availability exists, must match one.
        # If ANY negative availability exists, must NOT overlap.
        positive_avail = [a for a in availabilities if a.is_available]
        if positive_avail:
            fits = False
            s_str = start_time.strftime("%H:%M")
            e_str = end_time.strftime("%H:%M")
            for slot in positive_avail:
                if slot.start_time <= s_str and slot.end_time >= e_str:
                    fits = True
                    break
            if not fits:
                is_valid = False
                reasons.append("Outside available hours")
        
        negative_avail = [a for a in availabilities if not a.is_available]
        for slot in negative_avail:
            s_str = start_time.strftime("%H:%M")
            e_str = end_time.strftime("%H:%M")
            if s_str < slot.end_time and e_str > slot.start_time:
                is_valid = False
                reasons.append("During unavailable block")
                break
                
        if not is_valid:
            continue # Skip invalid candidates
            
        # 3. Weekly Hours Check
        start_of_week = start_time - timedelta(days=start_time.weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_week = start_of_week + timedelta(days=7)
        
        weekly_shifts = session.exec(select(Shift).where(
            Shift.employee_id == emp.id,
            Shift.start_time >= start_of_week,
            Shift.end_time < end_of_week
        )).all()
        
        current_weekly_hours = sum((s.end_time - s.start_time).total_seconds() / 3600 for s in weekly_shifts)
        new_shift_hours = (end_time - start_time).total_seconds() / 3600
        projected_weekly = current_weekly_hours + new_shift_hours
        
        if emp.max_weekly_hours and projected_weekly > emp.max_weekly_hours:
            score -= 50
            reasons.append(f"Overtime Risk: {projected_weekly:.1f} / {emp.max_weekly_hours} hrs")
        else:
            reasons.append(f"Weekly: {projected_weekly:.1f} hrs")
            
        # 4. Daily Hours Check
        # Get shifts for just this day
        start_of_day = start_time.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        daily_shifts = session.exec(select(Shift).where(
            Shift.employee_id == emp.id,
            Shift.start_time >= start_of_day,
            Shift.end_time < end_of_day
        )).all()
        
        current_daily_hours = sum((s.end_time - s.start_time).total_seconds() / 3600 for s in daily_shifts)
        projected_daily = current_daily_hours + new_shift_hours
        
        if projected_daily > 8:
            score -= 20
            reasons.append(f"Long Day: {projected_daily:.1f} hrs")
            
        recommendations.append({
            "employee": emp,
            "score": score,
            "reasons": reasons
        })
        
    # Sort by score desc
    # Sort by score desc, then by Hire Date asc (Seniority)
    # We use a tuple key: (-score, hire_date)
    # -score makes larger scores come first (since we sort ascending by default with this key)
    # hire_date asc makes older dates come first
    recommendations.sort(key=lambda x: (-x["score"], x["employee"].hire_date if x["employee"].hire_date else datetime.max))
    return recommendations

# --- Call Sheet Rotation ---
@app.get("/callsheet/rotation/")
def get_call_rotation(role_id: Optional[int] = None, session: Session = Depends(get_session)):
    """
    Returns employees sorted for call sheet rotation.
    Group 1: Full Time (Sorted by Hire Date)
    Group 2: Part Time + FT < Max Hours (Sorted by Hire Date)
    """
    query = select(Employee)
    if role_id:
        query = query.where(Employee.default_role_id == role_id)
        
    employees = session.exec(query).all()
    
    # Calculate weekly hours for FT employees to see if they should be on PT list
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_week = start_of_week + timedelta(days=7)
    
    full_time = []
    part_time = []
    
    for emp in employees:
        if emp.is_full_time:
            full_time.append(emp)
            
            # Check if under hours
            weekly_shifts = session.exec(select(Shift).where(
                Shift.employee_id == emp.id,
                Shift.start_time >= start_of_week,
                Shift.end_time < end_of_week
            )).all()
            
            hours = sum((s.end_time - s.start_time).total_seconds() / 3600 for s in weekly_shifts)
            max_hours = emp.max_weekly_hours if emp.max_weekly_hours else 40.0
            
            if hours < max_hours:
                # Add to PT list for extra shifts
                # We clone or just append? Appending is fine, frontend handles display.
                # Maybe add a flag or note?
                part_time.append(emp)
        else:
            part_time.append(emp)
    
    # Sort by Hire Date (Seniority = Oldest First)
    # Handle None hire_date (put at end)
    def sort_key(e):
        return e.hire_date if e.hire_date else datetime.max
        
    full_time.sort(key=sort_key)
    part_time.sort(key=sort_key)
    
    # Apply Rotation to Full Time List
    # "Start at the person after the last call"
    # 1. Find the employee with the most recent last_call_time
    last_called_emp = None
    most_recent_time = datetime.min
    
    for emp in full_time:
        if emp.last_call_time and emp.last_call_time > most_recent_time:
            most_recent_time = emp.last_call_time
            last_called_emp = emp
            
    if last_called_emp:
        try:
            # Find index in the seniority-sorted list
            # We use ID to match to avoid object identity issues if session refreshed
            idx = -1
            for i, e in enumerate(full_time):
                if e.id == last_called_emp.id:
                    idx = i
                    break
            
            if idx != -1:
                # Rotate: Elements after idx come first, then elements up to and including idx
                # List: [A, B, C, D] -> Last Called B (idx 1) -> [C, D, A, B]
                full_time = full_time[idx+1:] + full_time[:idx+1]
        except ValueError:
            pass # Should not happen given logic above

    return {
        "full_time": full_time,
        "part_time": part_time
    }

@app.post("/employees/{employee_id}/called/")
def update_last_call(employee_id: int, session: Session = Depends(get_session)):
    employee = session.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    employee.last_call_time = datetime.now()
    session.add(employee)
    session.commit()
    session.refresh(employee)
    return employee

# --- Authentication ---
class LoginRequest(BaseModel):
    password: str

@app.post("/login/")
def login(request: LoginRequest):
    # Simple password check for MVP
    # In production, use hashed passwords and env vars
    CORRECT_PASSWORD = "admin" 
    
    if request.password == CORRECT_PASSWORD:
        return {"success": True, "token": "fake-jwt-token"}
    if request.password == CORRECT_PASSWORD:
        return {"success": True, "token": "fake-jwt-token"}
    else:
        raise HTTPException(status_code=401, detail="Invalid password")

# --- Excel Export ---
from fastapi.responses import StreamingResponse

@app.get("/export/excel/")
def export_excel(session: Session = Depends(get_session)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Headers
    headers = ["Employee", "Role", "Date", "Start Time", "End Time", "Notes"]
    ws.append(headers)
    
    # Data
    shifts = session.exec(select(Shift).order_by(Shift.start_time)).all()
    
    for shift in shifts:
        emp_name = f"{shift.employee.first_name} {shift.employee.last_name}" if shift.employee else "OPEN"
        role_name = shift.role.name if shift.role else "Unknown"
        date_str = shift.start_time.strftime("%Y-%m-%d")
        start_str = shift.start_time.strftime("%H:%M")
        end_str = shift.end_time.strftime("%H:%M")
        
        ws.append([emp_name, role_name, date_str, start_str, end_str, shift.notes or ""])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="schedule_export.xlsx"'
    }
    headers = {
        'Content-Disposition': 'attachment; filename="schedule_export.xlsx"'
    }
    return StreamingResponse(buffer, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# --- OCR Import ---
import pytesseract
from PIL import Image
import io
from pdf2image import convert_from_bytes
import cv2
import numpy as np
import pillow_heif
import easyocr
import gc

# Register HEIF opener
pillow_heif.register_heif_opener()

# Initialize EasyOCR Reader (loads model into memory once)
# gpu=False to be safe, or True if available. False is safer for general compatibility.
reader = easyocr.Reader(['en'], gpu=False)

def deskew_image(image):
    # Convert PIL to OpenCV
    img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
    
    # Grayscale
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    
    # Invert (text is usually black on white, we want white on black for contours)
    gray = cv2.bitwise_not(gray)
    
    # Threshold to get text
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # Find all coordinates of non-zero pixels
    coords = np.column_stack(np.where(thresh > 0))
    
    # Find minimum area rectangle
    angle = cv2.minAreaRect(coords)[-1]
    
    # Correct angle
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
        
    # Rotate
    (h, w) = img_cv.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(img_cv, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    
    # Convert back to PIL
    return Image.fromarray(cv2.cvtColor(rotated, cv2.COLOR_BGR2RGB))

def correct_orientation(image):
    try:
        osd = pytesseract.image_to_osd(image)
        # More robust parsing - check if rotation info exists
        if '\nRotation: ' in osd:
            rotation_line = osd.split('\nRotation: ')[1].split('\n')[0]
            rotation = int(rotation_line)
            if rotation != 0:
                image = image.rotate(-rotation, expand=True)
                print(f"Rotated image by {-rotation} degrees")
        else:
            print("No rotation info in OSD, using original orientation")
            
    except Exception as e:
        print(f"Orientation detection failed: {e}, using original orientation")
        # Return original image on any error
        pass
    return image

def preprocess_image(image):
    # Convert to grayscale
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # Check dimensions
    height, width = gray.shape
    
    # Only rescale if image is small (e.g. < 2000px width)
    # If it's huge (e.g. 4000px+), downscale or keep as is
    if width < 2000:
        gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    elif width > 4000:
        # Downscale slightly to speed up processing without losing much detail for OCR
        gray = cv2.resize(gray, None, fx=0.5, fy=0.5, interpolation=cv2.INTER_AREA)
    
    # Apply Otsu's thresholding to binarize
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Denoise - reduce strength for speed on large images
    denoised = cv2.fastNlMeansDenoising(thresh, None, 10, 7, 21)
    
    return Image.fromarray(denoised)

@app.post("/import/ocr/")
async def import_ocr(dry_run: bool = False, file: UploadFile = File(...), session: Session = Depends(get_session)):
    print(f"OCR Request Received: {file.filename}, dry_run={dry_run}")
    try:
        contents = await file.read()
        
        # Initialize result containers
        parsed_shifts = []
        errors = []
        unmatched_lines = []
        unmatched_employees = []
        imported_count = 0
        
        # Convert PDF to image if needed
        images = []
        extracted_text = ""
        
        # 1. Try Direct Text Extraction for PDFs
        if file.filename.lower().endswith('.pdf'):
            try:
                from pypdf import PdfReader
                pdf_reader = PdfReader(io.BytesIO(contents))
                raw_text = ""
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    if text:
                        raw_text += text + "\n"
                
                if len(raw_text.strip()) > 50:
                    print("Direct PDF text extraction successful. Skipping OCR.")
                    extracted_text = raw_text
                else:
                    print("PDF has insufficient text (likely scanned). Falling back to OCR.")
                    raise Exception("Insufficient text")
            except Exception as e:
                print(f"Direct text extraction skipped: {e}")
                # Fallback to Image Extraction
                try:
                    images = convert_from_bytes(contents)
                except Exception as e:
                    print(f"pdf2image failed (likely missing poppler): {e}")
                    # Fallback: Try extracting images with pypdf
                    try:
                        from pypdf import PdfReader
                        pdf_reader = PdfReader(io.BytesIO(contents))
                        for page in pdf_reader.pages:
                            for image_file_object in page.images:
                                images.append(Image.open(io.BytesIO(image_file_object.data)))
                        
                        if not images:
                            raise Exception("No images found in PDF (and poppler is missing for rendering text PDFs).")
                        print(f"Successfully extracted {len(images)} images via pypdf fallback.")
                    except Exception as pypdf_error:
                        print(f"pypdf fallback failed: {pypdf_error}")
                        raise Exception("PDF processing failed. Please install 'poppler' (brew install poppler) or upload an image.")
        else:
            images = [Image.open(io.BytesIO(contents))]
            
        # Ensure all images are RGB for OpenCV/EasyOCR compatibility
        if images:
            images = [img.convert('RGB') for img in images]
            
        # Only run OCR if we don't have text yet
        extracted_text = ""
        determined_angle = None
        column_dates = {}
        
        for img in images:
            # 1. Correct Orientation (90/180/270)
            # We still run this fast check as it might catch simple flips
            img = correct_orientation(img)

            # EasyOCR Strategy
            # Convert PIL to bytes or numpy array for EasyOCR
            img_np = np.array(img)
            
            results = []
            
            # 4-Way Rotation Check
            # If we haven't determined the angle yet (first page), run the check
            if determined_angle is None:
                best_results = []
                best_score = -1
                best_angle = 0
                
                import re
                
                for angle in [0, 90, 180, 270]:
                    # Rotate image
                    rotated_img = img.rotate(-angle, expand=True)
                    img_np_rot = np.array(rotated_img)
                    
                    # Run EasyOCR
                    curr_results = reader.readtext(img_np_rot, detail=1, paragraph=False, x_ths=0.5)
                    
                    # Score this orientation
                    score = 0
                    text_content = " ".join([r[1] for r in curr_results])
                    
                    # Check for time patterns (e.g. 9:00, 9-5)
                    time_matches = re.findall(r'\d{1,2}[:\.]?\d{0,2}\s*-\s*\d{1,2}[:\.]?\d{0,2}', text_content)
                    score += len(time_matches) * 2
                    
                    # Check for day names
                    days = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
                    day_matches = [d for d in text_content.lower().split() if any(day in d for day in days)]
                    score += len(day_matches)
                    
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"Angle {angle}: Score {score} (Times: {len(time_matches)}, Days: {len(day_matches)})\n")
                    
                    if score > best_score:
                        best_score = score
                        best_results = curr_results
                        best_angle = angle
                
                determined_angle = best_angle
                results = best_results
                
                with open("ocr_debug.log", "a") as f:
                    f.write(f"Determined Document Angle: {determined_angle} with Score {best_score}\n")
            else:
                # Use determined angle for subsequent pages
                if determined_angle != 0:
                    img = img.rotate(-determined_angle, expand=True)
                    img_np = np.array(img)
                
                # Use x_ths=0.5 to prevent merging of close words (like headers)
                results = reader.readtext(img_np, detail=1, paragraph=False, x_ths=0.5)
        
            # Sort by Y
            results.sort(key=lambda x: x[0][0][1])
        
        # Group into lines
            lines_data = [] # List of lists of (bbox, text)
            current_line = []
            last_y = -1
            
            for (bbox, text, prob) in results:
                y = bbox[0][1]
                if last_y == -1:
                    current_line.append((bbox, text))
                    last_y = y
                    continue
                
                if abs(y - last_y) < 35:
                    current_line.append((bbox, text))
                else:
                    current_line.sort(key=lambda x: x[0][0][0])
                    lines_data.append(current_line)
                    current_line = [(bbox, text)]
                    last_y = y
            if current_line:
                current_line.sort(key=lambda x: x[0][0][0])
                lines_data.append(current_line)
                
            # Process Lines for this Page
            
            # --- DEBUG: Save Raw Data for Offline Parsing ---
            import json
            # Convert numpy types to python types for JSON serialization if needed
            # EasyOCR output is usually list of (bbox, text, prob)
            # bbox is list of [x,y] points.
            
            # Helper to serialize
            def serialize_ocr_data(data):
                serializable = []
                for line in data:
                    ser_line = []
                    for (bbox, text) in line:
                        # bbox is list of 4 points [[x,y], [x,y], [x,y], [x,y]]
                        # Convert to list of lists (if it's not already)
                        ser_bbox = [[float(p[0]), float(p[1])] for p in bbox]
                        ser_line.append((ser_bbox, text))
                    serializable.append(ser_line)
                return serializable

            try:
                with open("ocr_raw_output.json", "w") as f:
                    json.dump(serialize_ocr_data(lines_data), f, indent=2)
                print("Saved raw OCR data to ocr_raw_output.json")
            except Exception as e:
                print(f"Failed to save raw debug data: {e}")
            # -----------------------------------------------

            day_columns = [] # List of (x_center, day_index)
            # column_dates = {} # Map col_idx -> datetime -- MOVED OUTSIDE LOOP TO PERSIST
            pending_date_row = None # Store date row if found before header
            header_y = -1
            current_location = "General" # Default
            
            # PREPROCESSING: Merge split rows (e.g., time slots on one row, name on next row)
            merged_lines = []
            i = 0
            while i < len(lines_data):
                current_row = lines_data[i]
                current_text = " ".join([t[1] for t in current_row])
                
                # Check if next row exists
                if i + 1 < len(lines_data):
                    next_row = lines_data[i + 1]
                    next_text = " ".join([t[1] for t in next_row])
                    
                    # Pattern: Current row has times but starts with just a number or short text
                    # Next row looks like it has a name
                    # Look for time patterns in current row
                    time_pattern = r'\d{1,2}[:.,]\d{2}\s*[APap]'
                    has_times_current = len(re.findall(time_pattern, current_text)) >= 2
                    
                    # Check if current row starts with just a number (like "3")
                    first_word = current_text.strip().split()[0] if current_text.strip() else ""
                    starts_with_number = len(first_word) <= 2 and first_word.isdigit()
                    
                    # Check if next row starts with a name-like pattern (letters, possibly with periods/spaces)
                    next_first_words = next_text.strip().split()[:2]
                    looks_like_name = (len(next_first_words) >= 1 and 
                                     any(c.isalpha() for c in next_first_words[0]) and
                                     len(next_first_words[0]) > 2)
                    
                    # If current has times but weak name, and next has strong name, merge
                    if has_times_current and starts_with_number and looks_like_name:
                        # Merge: Combine items from both rows
                        merged_row = next_row + current_row  # Put name first, then time slots
                        merged_lines.append(merged_row)
                        i += 2  # Skip both rows
                        continue
                
                # No merge, just add current row
                merged_lines.append(current_row)
                i += 1
            
            lines_data = merged_lines  # Use merged data
            
            for line_items in lines_data:
                # Construct full text for regex checks
                full_line_text = " ".join([t[1] for t in line_items])
                
                with open("ocr_debug.log", "a") as f:
                    f.write(f"DEBUG: Raw Line: {full_line_text}\n")
                
                # 0. Check Location
                line_upper = full_line_text.upper()
                import re # Ensure re is available
                
                for loc in KNOWN_LOCATIONS + list(LOCATION_MAPPINGS.keys()):
                    # Use regex to ensure we don't match "Lot 2" inside "Lot 2:45"
                    # We look for the location string, NOT followed by a time separator (: or .)
                    # We also want to match "C-Lot" which might be "C-LOT" or "CLOT"
                    
                    # Escape the location string for regex
                    loc_pattern = re.escape(loc)
                    
                    # Regex: Match loc with word boundaries to avoid partial matches
                    # e.g. "AD" should not match "(Ad)" or "Add"
                    # e.g. "LOT 2" should not match "LOT 2:45"
                    # We use \b for word boundaries, but we also need to handle the case where
                    # the location might be at the start/end of the string or surrounded by non-word chars like ()
                    # However, \b matches between \w and \W. 
                    # "LOT 2" has a space, so \bLOT 2\b works for " LOT 2 "
                    # But "AD" in "(AD)" -> "(" is \W, "A" is \w, so \b matches before A.
                    line_upper = full_line_text.upper()
                    
                    # First: Check if this is a VERY short line with just the location (section header)
                    # This is most likely a page section header
                    is_section_header = False
                    if len(full_line_text.strip()) < 25:  # Very short line
                        # Check if the entire line is basically just the location name
                        clean_line = re.sub(r'[^A-Z0-9\s-]', '', line_upper).strip()
                        if loc in clean_line or clean_line in loc:
                            is_section_header = True
                    
                    # Second: Original logic for slightly longer lines
                    is_header_candidate = False
                    if not is_section_header and len(full_line_text) < 40:
                        is_header_candidate = True
                    
                    pattern = rf"\b{loc_pattern}\b(?!\s*[:.;])"
                    match = re.search(pattern, line_upper)
                    
                    # Process if it's a section header OR (candidate AND matches pattern)
                    if is_section_header or (match and is_header_candidate):
                        if loc in LOCATION_MAPPINGS:
                            current_location = LOCATION_MAPPINGS[loc]
                        else:
                            current_location = loc.title()
                            if loc == "CONRAC": current_location = "Conrac"
                            if loc == "PLAZA": current_location = "Plaza"
                            if loc == "LOT 1": current_location = "Lot 1"
                            if loc == "LOT 2": current_location = "Lot 2"
                            if loc == "LOT 3": current_location = "Lot 3"
                            if loc == "LOT 4": current_location = "Lot 4"
                        
                        # Special case: Supervisors don't need a location
                        if loc == "SUPERVISORS" or current_location == "Supervisors":
                            current_location = None
                        
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Found Location Header ({'SECTION' if is_section_header else 'STRICT'}): {current_location or 'None (Supervisors)'} in line: {full_line_text}\n")
                        break
                    
                # 1. Check Header (Define Columns)
                with open("ocr_debug.log", "a") as f:
                    f.write(f"DEBUG: Finished Location Check for: {full_line_text[:30]}...\n")
                
                days = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
                day_matches = []
                
                for (bbox, text) in line_items:
                    # Check if this text block contains day names
                    found_days = [d for d in days if d in text.lower()]
                    
                    # DEBUG: Print EVERYTHING
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Checking Block: '{text}' (Type: {type(text)}) -> Found: {found_days}\n")

                    if found_days:
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Header Candidate Block: '{text}' -> Found Days: {found_days}\n")
                        
                        if len(found_days) == 1:
                            # Single day in this block
                            x_center = (bbox[0][0] + bbox[1][0]) / 2
                            day_matches.append((x_center, text))
                        else:
                            # Multiple days merged in one block (e.g. "SATURDAY SUNDAY" or "SATURDAYSUNDAY")
                            words = text.split()
                            
                            # If we have fewer words than found days, it means some are glued together without spaces
                            if len(words) < len(found_days):
                                # Distribute based on found_days count
                                box_width = bbox[1][0] - bbox[0][0]
                                start_x = bbox[0][0]
                                segment_width = box_width / len(found_days)
                                
                                for i, day_name in enumerate(found_days):
                                    # Estimate center for this segment
                                    seg_center = start_x + (segment_width * i) + (segment_width / 2)
                                    day_matches.append((seg_center, day_name))
                            else:
                                # Words are separated by space, but in one block? 
                                # EasyOCR usually splits spaces, but if not:
                                # We can try to map words to days?
                                # For now, let's just use the block center for all (bad) or distribute
                                # Distributing is safer
                                box_width = bbox[1][0] - bbox[0][0]
                                start_x = bbox[0][0]
                                segment_width = box_width / len(found_days)
                                for i, day_name in enumerate(found_days):
                                    seg_center = start_x + (segment_width * i) + (segment_width / 2)
                                    day_matches.append((seg_center, day_name))
                
                if len(day_matches) >= 3:
                    # Found header! Define columns
                    day_columns = sorted(day_matches, key=lambda x: x[0])
                    header_y = line_items[0][0][0][1] # Y of header
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Defined {len(day_columns)} columns based on header: {[d[1].upper() for d in day_columns]}\n")
                    
                    # Check for pending date row
                    if pending_date_row:
                        for (d_x, d_str) in pending_date_row:
                            # Find nearest column
                            closest_col_idx = -1
                            min_dist = float('inf')
                            
                            for i, (col_x, col_name) in enumerate(day_columns):
                                dist = abs(d_x - col_x)
                                if dist < min_dist:
                                    min_dist = dist
                                    closest_col_idx = i
                            
                            if min_dist < 150: # Threshold
                                column_dates[closest_col_idx] = d_str
                        
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Applied pending date row to columns: {column_dates}\n")
                        pending_date_row = None # Clear it
                    continue
                
                # 1.5 Check for Date Row (e.g. 12/6/25)
                # Check for dates regardless of whether we have columns yet
                # If we find them before columns, store them as pending
                date_matches = []
                date_regex = r'\d{1,6}/\d{1,6}(?:/\d{2,4})?'  # More lenient to catch OCR errors
                for (bbox, text) in line_items:
                    found_dates = re.findall(date_regex, text)
                    if found_dates:
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Date Regex Match in '{text}': {found_dates}\n")
                        # Clean and normalize dates
                        cleaned_dates = []
                        for d in found_dates:
                            with open("ocr_debug.log", "a") as f:
                                f.write(f"DEBUG: Cleaning date '{d}'. Count(/): {d.count('/')}\n")
                            
                            # Try to fix common OCR errors like '12/6125' -> '12/6/25'
                            # or '12111/25' -> '12/11/25'
                            original_d = d
                            if d.count('/') == 1:
                                parts = d.split('/')
                                # Case 1: Missing 2nd slash (12/6125)
                                if len(parts[1]) >= 4:  # e.g., '6125' or '7125'
                                    day_year = parts[1]
                                    if len(day_year) == 4:  # e.g., '6125'
                                        day = day_year[:1]
                                        year = day_year[1:]
                                        if len(year) == 3 and year.startswith('1'):
                                            year = year[1:]
                                        d = f"{parts[0]}/{day}/{year}"
                                    elif len(day_year) == 5:  # e.g., '12125'
                                        day = day_year[:2]
                                        year = day_year[2:]
                                        if len(year) == 3 and year.startswith('1'):
                                            year = year[1:]
                                        d = f"{parts[0]}/{day}/{year}"
                                # Case 2: Missing 1st slash (12111/25 -> 12/11/25)
                                elif len(parts[0]) >= 3:
                                    # e.g. 1219 -> 12/9, 12110 -> 12/10
                                    # Assume first 2 digits are month, next digit is separator (often '1'), rest is day
                                    p0 = parts[0]
                                    if len(p0) >= 3:
                                        month = p0[:2]
                                        rest = p0[2:]
                                        # If rest starts with '1' and has more digits, assume '1' is slash
                                        if rest.startswith('1') and len(rest) > 1:
                                            day = rest[1:]
                                            d = f"{month}/{day}/{parts[1]}"
                                        # Fallback: just split? 1219 -> 12/9?
                                        elif len(rest) >= 1:
                                            day = rest
                                            year = parts[1]
                                            d = f"{month}/{day}/{year}"
                            
                            elif d.count('/') == 2:
                                parts = d.split('/')
                                if len(parts) == 3:
                                    year = parts[2]
                                    with open("ocr_debug.log", "a") as f:
                                        f.write(f"DEBUG: Checking year '{year}' in '{d}'. Len: {len(year)}\n")
                                    
                                    # Case 3: Year has 3 digits (e.g. 125 -> 25)
                                    if len(year) == 3 and year.startswith('1'):
                                        year = year[1:]
                                        d = f"{parts[0]}/{parts[1]}/{year}"
                            
                            if d != original_d:
                                with open("ocr_debug.log", "a") as f:
                                    f.write(f"DEBUG: Cleaned date '{original_d}' -> '{d}'\n")
                            
                            cleaned_dates.append(d)
                        
                        # If multiple dates in one block?
                        if len(cleaned_dates) == 1:
                            x_center = (bbox[0][0] + bbox[1][0]) / 2
                            date_matches.append((x_center, cleaned_dates[0]))
                        else:
                            # Distribute
                            box_width = bbox[1][0] - bbox[0][0]
                            start_x = bbox[0][0]
                            segment_width = box_width / len(cleaned_dates)
                            for i, d_str in enumerate(cleaned_dates):
                                seg_center = start_x + (segment_width * i) + (segment_width / 2)
                                date_matches.append((seg_center, d_str))
                
                if len(date_matches) >= 3:
                    # Found a date row!
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Found Date Row with {len(date_matches)} dates. day_columns defined? {bool(day_columns)}\n")
                    
                    if day_columns:
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Entering mapping block. Columns: {len(day_columns)}\n")
                        # Map immediately if we have columns
                        for (d_x, d_str) in date_matches:
                            # Find nearest column
                            closest_col_idx = -1
                            min_dist = float('inf')
                            
                            for i, (col_x, col_name) in enumerate(day_columns):
                                dist = abs(d_x - col_x)
                                if dist < min_dist:
                                    min_dist = dist
                                    closest_col_idx = i
                            
                            if min_dist < 150: # Threshold
                                column_dates[closest_col_idx] = d_str
                        
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Mapped dates to columns: {column_dates}\n")
                    else:
                        # Store for later
                        pending_date_row = date_matches
                        with open("ocr_debug.log", "a") as f:
                            f.write(f"DEBUG: Stored pending date row (waiting for columns)\n")
                    continue


                
                # 2. Process Row (if we have columns defined)
                if day_columns:
                    # We have columns, try to map items to them
                    # Identify Name (Leftmost) vs Time Slots
                    
                    # Heuristic: Name is usually far left, before the first column starts
                    # First column X
                    first_col_x = day_columns[0][0]
                    margin = 50 
                    
                    name_parts = []
                    time_slots = {} 
                    
                    # Calculate average column width for gap detection
                    col_xs = [d[0] for d in day_columns]
                    if len(col_xs) > 1:
                        avg_col_gap = (col_xs[-1] - col_xs[0]) / (len(col_xs) - 1)
                    else:
                        avg_col_gap = 200 # Fallback
                    
                    for (bbox, text) in line_items:
                        x_center = (bbox[0][0] + bbox[1][0]) / 2
                        width = bbox[1][0] - bbox[0][0]
                        
                        if x_center < (first_col_x - margin):
                            name_parts.append(text)
                        else:
                            # Check if block is wide (spans multiple columns)
                            if width > (avg_col_gap * 1.2):
                                # Merged block! Split it.
                                start_x = bbox[0][0]
                                end_x = bbox[1][0]
                                
                                covered_cols = []
                                for k, (col_x, col_name) in enumerate(day_columns):
                                    if start_x - (avg_col_gap/2) <= col_x <= end_x + (avg_col_gap/2):
                                        covered_cols.append(k)
                                
                                if covered_cols:
                                    words = text.split()
                                    # Distribute words
                                    if len(words) >= len(covered_cols):
                                        chunk_size = len(words) / len(covered_cols)
                                        for i, col_idx in enumerate(covered_cols):
                                            s = int(i * chunk_size)
                                            e = int((i + 1) * chunk_size)
                                            chunk = " ".join(words[s:e])
                                            if col_idx in time_slots:
                                                time_slots[col_idx] += " " + chunk
                                            else:
                                                time_slots[col_idx] = chunk
                                    else:
                                        # Not enough word breaks - evenly divide and extract patterns
                                        char_len = len(text)
                                        chunk_size = char_len // len(covered_cols)
                                        
                                        # Pattern to extract time strings (including OCR errors O/0 I/1 S/5)
                                        time_pattern = r'[0-9IO]{1,2}[:.]?[0-9IO]{0,2}[APMS]{0,3}[-–][0-9IO]{1,2}[:.]?[0-9IO]{0,2}[APMS]{0,3}'
                                        
                                        for i, col_idx in enumerate(covered_cols):
                                            start = i * chunk_size
                                            end = (i + 1) * chunk_size if i < len(covered_cols) - 1 else char_len
                                            segment = text[start:end]
                                            
                                            # Try to find a time pattern within this segment
                                            match = re.search(time_pattern, segment, re.IGNORECASE)
                                            if match:
                                                chunk = match.group()
                                            else:
                                                chunk = segment.strip()
                                            
                                            if col_idx in time_slots:
                                                time_slots[col_idx] += " " + chunk
                                            else:
                                                time_slots[col_idx] = chunk
                                else:
                                    # Fallback to center mapping
                                    closest_col_idx = -1
                                    min_dist = float('inf')
                                    for k, (col_x, col_name) in enumerate(day_columns):
                                        dist = abs(x_center - col_x)
                                        if dist < min_dist:
                                            min_dist = dist
                                            closest_col_idx = k
                                    if closest_col_idx != -1:
                                        if closest_col_idx in time_slots:
                                            time_slots[closest_col_idx] += " " + text
                                        else:
                                            time_slots[closest_col_idx] = text
                            else:
                                # Normal mapping (not wide)
                                closest_col_idx = -1
                                min_dist = float('inf')
                                for k, (col_x, col_name) in enumerate(day_columns):
                                    dist = abs(x_center - col_x)
                                    if dist < min_dist:
                                        min_dist = dist
                                        closest_col_idx = k
                                
                                if closest_col_idx != -1:
                                    # Check distance
                                    if min_dist < (avg_col_gap * 0.6):
                                        if closest_col_idx in time_slots:
                                            time_slots[closest_col_idx] += " " + text
                                        else:
                                            time_slots[closest_col_idx] = text

                    full_name = " ".join(name_parts).replace('_', ' ').strip()
                    # Remove trailing dots/chars
                    full_name = re.sub(r'[.:,]+$', '', full_name).strip()
                    
                    if not full_name: continue
                    
                    # Match Employee - Try full name first, then first name only
                    name_parts_clean = full_name.split()
                    first_word = name_parts_clean[0] if name_parts_clean else ""
                    
                    employee = None
                    # Try to match with last name if we have 2+ words
                    if len(name_parts_clean) >= 2:
                        # Try exact match on first + last
                        for emp in session.exec(select(Employee)).all():
                            emp_full = f"{emp.first_name} {emp.last_name}".lower()
                            # Check if OCR name contains employee's last name
                            if name_parts_clean[0].lower() == emp.first_name.lower():
                                # First name matches, check if last name is in OCR text
                                for part in name_parts_clean[1:]:
                                    if part.lower() in emp.last_name.lower() or emp.last_name.lower() in part.lower():
                                        employee = emp
                                        break
                            if employee:
                                break
                    
                    # Fallback to first name only if no match found
                    if not employee:
                        employee = session.exec(select(Employee).where(Employee.first_name.ilike(first_word))).first()
                    
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Processing Row: '{full_name}' -> Employee Match: {employee.first_name if employee else 'None'} (Loc: {current_location})\n")
                    
                    # Define helper once
                    def parse_ocr_time(t_str):
                        # OCR Error Correction BEFORE parsing
                        t_str = t_str.strip().upper()
                        
                        # Common OCR mistakes
                        t_str = t_str.replace('S', '5')  # 9:4SAM → 9:45AM
                        t_str = t_str.replace('O', '0')  # 1O:15 → 10:15
                        t_str = t_str.replace(';', ':')  # 11;15 → 11:15
                        t_str = t_str.replace('.', ':')  # 9.45 → 9:45
                        # Smart space handling: convert "5 0P" to "5:0P" before removing spaces
                        # This prevents "5 0P" → "50P" (hour=50 error)
                        t_str = re.sub(r'(\d)\s+(\d)', r'\1:\2', t_str)  # digit-space-digit → digit:digit
                        t_str = t_str.replace(' ', '')   # Now safe to remove remaining spaces
                        t_str = t_str.replace('I', '1')  # I0:15 → 10:15
                        t_str = t_str.replace('L', '1')  # L:45 → 1:45
                        
                        is_pm = 'P' in t_str
                        is_am = 'A' in t_str
                        # Remove letters
                        t_str = re.sub(r'[A-Z]', '', t_str)
                        
                        h = 0
                        m = 0
                        if ':' in t_str:
                            parts = t_str.split(':')
                            h = int(parts[0])
                            m = int(parts[1])
                        elif len(t_str) >= 3:
                            # 3 or 4 digits: 930, 1030
                            m = int(t_str[-2:])
                            h = int(t_str[:-2])
                        else:
                            # 1 or 2 digits: 9, 10
                            h = int(t_str)
                            
                        # PM Logic
                        if is_pm and h != 12: h += 12
                        elif is_am and h == 12: h = 0
                        elif not is_pm and not is_am:
                            if h < 7: h += 12
                            
                        return time(hour=h, minute=m)

                    # Log time slots for debugging
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Row '{full_name}' Time Slots: {time_slots}\n")
                        f.write(f"DEBUG: Day Columns (X): {[(c[0], c[1]) for c in day_columns]}\n")

                    # Iterate columns to find shifts
                    row_shifts = []
                    
                    # Process each column to find shifts
                    for col_idx, (col_x, col_text) in enumerate(day_columns):
                        try:

                            # Determine day offset
                            day_lower = col_text.lower()
                            day_offset = 0
                            if "mon" in day_lower: day_offset = 0
                            elif "tue" in day_lower: day_offset = 1
                            elif "wed" in day_lower: day_offset = 2
                            elif "thu" in day_lower: day_offset = 3
                            elif "fri" in day_lower: day_offset = 4
                            elif "sat" in day_lower: day_offset = 5
                            elif "sun" in day_lower: day_offset = 6
                            
                            # Get time text for this column
                            time_text = time_slots.get(col_idx, "OFF") 
                            
                            # Skip if OFF

                            if "off" in time_text.lower():
                                continue
                                
                            # Check for location override in the text (Use raw text)
                            shift_location = current_location
                            text_upper_raw = time_text.upper()
                            
                            # Check mappings first
                            found_override = False
                            for key, val in LOCATION_MAPPINGS.items():
                                if key in text_upper_raw:
                                    shift_location = val
                                    found_override = True
                                    break
                            
                            if not found_override:
                                for loc in KNOWN_LOCATIONS:
                                    if loc in text_upper_raw:
                                        shift_location = loc.title()
                                        if loc == "CONRAC": shift_location = "Conrac"
                                        if loc == "PLAZA": shift_location = "Plaza"
                                        break

                            # Clean common OCR typos first
                            clean_time_text = time_text.upper()
                            
                            # Filter out known non-time text (Locations/OFF) - Expanded list
                            if any(x in clean_time_text for x in ['LOT', 'PLAZA', 'CONRAC', 'OFF', 'VACATION', '10T', 'P1AZA', 'P1A2A', 'C-LOT', 'E-LOT']):
                                continue

                            # Step 1: Fix merged colons FIRST before any other transformations
                            # Patterns like "12004" (5 digits) should become "12:00", "1200" (4 digits) -> "12:00", "830A" -> "8:30A"
                            # OCR often reads "12:00" as "12004" where the colon becomes an extra '0'
                            # First, handle 5-digit patterns (e.g., 12004 -> 12:00)
                            # Match: 1-2 digits, then 2-3 more digits (one may be extra), followed by A/P or dash/space
                            clean_time_text = time_text.upper()
                            # Fix 5-digit times like "12004" -> "12:00" (take first 1-2 digits, then next 2, ignore extra)
                            clean_time_text = re.sub(r'\b(\d{1,2})(\d{2})\d?(?=[APap\-\s]|$)', r'\1:\2', clean_time_text)
                            
                            # Step 2: Character replacements for OCR errors
                            clean_time_text = clean_time_text.replace('O', '0').replace('Q', '0').replace('D', '0')
                            clean_time_text = clean_time_text.replace('I', '1').replace('L', '1').replace('!', '1').replace('|', '1')
                            # Fix common time OCR pattern: "10BAM" or "1OBAM" should be "10:00AM"
                            # B in this context is a misread '0', not '8'
                            clean_time_text = re.sub(r'\b(1[0-2]?)B([AP])', r'\1:00\2', clean_time_text, flags=re.IGNORECASE)
                            clean_time_text = clean_time_text.replace('B', '8').replace('S', '5').replace('Z', '2')
                            # Treat underscores as hyphens (range separators), not spaces
                            clean_time_text = clean_time_text.replace('_', '-').replace('.', ':').replace(';', ':').replace(',', ':')
                            
                            # Step 3: Fix jammed times (e.g., "1200A8.30A" -> "1200A-8.30A")
                            # Pattern: time with A/P directly followed by another time
                            # Match: digits:digits followed by A/P, then immediately digits
                            clean_time_text = re.sub(r'([AP])(\d{1,2}[:.]?\d{0,2}[AP])', r'\1-\2', clean_time_text, flags=re.IGNORECASE)
                            
                            # Step 4: Fix missing hyphen if numbers are jammed (e.g. "6:00P2:00A")
                            clean_time_text = clean_time_text.replace('-', ' - ')
                            
                            # Step 4: Infer missing AM/PM markers in time ranges
                            # E.g., "12:00 - 8:30A" should become "12:00A - 8:30A"
                            # Pattern: time without A/P, dash, time with A/P
                            # Match: digits:digits (no A/P), then dash, then digits:digits followed by A/P
                            clean_time_text = re.sub(r'(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})([AP])', r'\1\3 - \2\3', clean_time_text, flags=re.IGNORECASE)
                            
                            # Regex pattern allowing flexible separators (colon, period, comma, semicolon, space, or none)
                            time_part_regex = r'(?:\d{1,2}[:.,;\s]?\d{0,2})\s*[APap][Mm]?'
                            range_pattern_regex = f'({time_part_regex}\\s*-\\s*{time_part_regex})'
                            
                            time_matches_in_slot = re.findall(range_pattern_regex, clean_time_text, re.IGNORECASE)
                            
                            with open("ocr_debug.log", "a") as f:
                                f.write(f"DEBUG: Col {col_idx} Original='{time_text}', Cleaned='{clean_time_text}', Matches={time_matches_in_slot}\n")

                            # Calculate base date for this week
                            # IMPROVEMENT: Use an "Anchor Date" from column_dates if available, 
                            # instead of defaulting to datetime.now()
                            start_of_week = None
                            
                            # Try to find an anchor date
                            if column_dates:
                                for c_idx, d_str in column_dates.items():
                                    try:
                                        # Parse the date string
                                        # d_str format is likely MM/DD/YY or MM/DD/YYYY from our regex/cleaning
                                        # We need to handle 2-digit years
                                        parts = d_str.split('/')
                                        if len(parts) == 3:
                                            m, d, y = map(int, parts)
                                            if y < 100: y += 2000 # Assume 20xx
                                            anchor_date = datetime(y, m, d)
                                            
                                            # Determine offset of this column
                                            # We need to know which day of week this column corresponds to
                                            # We have day_columns[c_idx] -> (x, "MONDAY")
                                            if c_idx < len(day_columns):
                                                col_name = day_columns[c_idx][1].lower()
                                                anchor_offset = 0
                                                if "mon" in col_name: anchor_offset = 0
                                                elif "tue" in col_name: anchor_offset = 1
                                                elif "wed" in col_name: anchor_offset = 2
                                                elif "thu" in col_name: anchor_offset = 3
                                                elif "fri" in col_name: anchor_offset = 4
                                                elif "sat" in col_name: anchor_offset = 5 # Sat is usually before Sun in this schedule?
                                                elif "sun" in col_name: anchor_offset = 6
                                                
                                                # If schedule is Sat-Fri, we need to be careful about "start of week"
                                                # Let's assume standard Mon-Sun week for start_of_week calculation
                                                # If Sat 12/6 is anchor (offset 5), then Mon 12/1 is start of THAT week?
                                                # Or is Sat 12/6 part of the week starting Mon 12/8?
                                                # The user said: Sat 12/6, Sun 12/7, Mon 12/8...
                                                # So Sat/Sun are the *previous* week's weekend, or the schedule starts on Saturday?
                                                # If Sat 12/6 is "Day 0" of the schedule...
                                                
                                                # Let's just calculate the date for the CURRENT column relative to the anchor
                                                # We don't strictly need "start_of_week" if we map offsets relative to anchor
                                                pass

                                            # Calculate Monday of the week containing the anchor
                                            # But wait, Sat 12/6 and Mon 12/8 are in the SAME row.
                                            # Standard ISO week: Mon 12/1 -> Sun 12/7. Mon 12/8 is next week.
                                            # So Sat 12/6 is in week 1, Mon 12/8 is in week 2.
                                            
                                            # Let's rely on the column headers to define the relative structure
                                            # If we have Sat, Sun, Mon, Tue...
                                            # And we have a date for Sat (12/6).
                                            # We want to find the date for Mon.
                                            # Mon is 2 columns after Sat? No, Sat(0), Sun(1), Mon(2).
                                            # So Mon is +2 days from Sat?
                                            # 12/6 + 2 days = 12/8. CORRECT.
                                            
                                            # So we can just use the anchor date and the difference in column indices?
                                            # ONLY IF columns are consecutive days.
                                            # Let's assume they are roughly consecutive or use the day names.
                                            
                                            # Better approach:
                                            # 1. Identify the day-of-week index (0=Mon, 6=Sun) for the ANCHOR.
                                            # 2. Identify the day-of-week index for the TARGET column.
                                            # 3. Calculate difference.
                                            # 4. Handle wrap-around? 
                                            #    Sat(5) -> Sun(6) -> Mon(0). 
                                            #    Difference: 6-5=1 (Sun is +1 day). 
                                            #    0-5 = -5? No, Mon is AFTER Sun.
                                            #    If the schedule is Sat, Sun, Mon... then Mon is +2 days from Sat.
                                            
                                            # Let's use the list index in day_columns as the truth for "days from start of row"
                                            # If day_columns is [Sat, Sun, Mon, Tue...]
                                            # Anchor = Sat (idx 0) = 12/6
                                            # Target = Mon (idx 2)
                                            # Target Date = Anchor Date + (Target Idx - Anchor Idx) days
                                            # 12/6 + (2-0) = 12/8. PERFECT.
                                            
                                            start_of_week = anchor_date - timedelta(days=c_idx) # Virtual start date (Day 0 of the row)
                                            break
                                    except:
                                        continue
                            
                            if not start_of_week:
                                # Fallback to system time if NO dates found in the entire row
                                today = datetime.now()
                                # Default to Monday of current week? Or just today?
                                # Let's stick to Monday of current week as a safe default
                                start_of_week = today - timedelta(days=today.weekday())

                            # Determine specific date for this shift
                            # Determine specific date for this shift
                            if col_idx in column_dates:
                                # Use explicit date if found
                                try:
                                    val = column_dates[col_idx]
                                    if isinstance(val, datetime):
                                        current_shift_date = val
                                    elif isinstance(val, str):
                                        parts = val.split('/')
                                        if len(parts) == 3:
                                            m, d, y = map(int, parts)
                                            if y < 100: y += 2000
                                            current_shift_date = datetime(y, m, d)
                                        else:
                                            current_shift_date = start_of_week + timedelta(days=col_idx)
                                    else:
                                        current_shift_date = start_of_week + timedelta(days=col_idx)
                                except:
                                    current_shift_date = start_of_week + timedelta(days=col_idx)
                            else:
                                # Calculate based on virtual start of row
                                current_shift_date = start_of_week + timedelta(days=col_idx)

                            if not time_matches_in_slot:
                                # Fallback: Store raw text if meaningful
                                # if len(clean_time_text) > 3:
                                #     # Create a dummy shift with raw text in notes
                                #     s_dt = current_shift_date.replace(hour=9, minute=0, second=0, microsecond=0)
                                #     e_dt = current_shift_date.replace(hour=17, minute=0, second=0, microsecond=0)
                                #     
                                #     row_shifts.append({
                                #         "start_time": s_dt,
                                #         "end_time": e_dt,
                                #         "location": shift_location,
                                #         "notes": f"RAW: {clean_time_text}" # Flag for frontend
                                #     })
                                #     with open("ocr_debug.log", "a") as f:
                                #         f.write(f"DEBUG: Appended Fallback Shift for Col {col_idx}\n")
                                continue 
                            
                            match_str = time_matches_in_slot[0] 
                            
                            # Parse Time Range - Strip spaces around dash first
                            match_str = match_str.replace(' - ', '-').replace('- ', '-').replace(' -', '-')
                            # Replace period/comma/semicolon with colon (but NOT spaces)
                            match_str = re.sub(r'[.,;]', ':', match_str)
                            t_parts = match_str.split('-')
                            if len(t_parts) != 2: continue
                            
                            t_start_str, t_end_str = t_parts
                            
                            try:
                                s_time = parse_ocr_time(t_start_str)
                                e_time = parse_ocr_time(t_end_str)
                                
                                # Handle overnight shifts (end < start)
                                # Note: e_time is a time object, we track overnight by comparing times
                                # If end < start, we already added 1 day above, so end_dt calculation below handles it
                                if e_time < s_time:
                                    # Overnight shift detected - end_dt needs to be next day
                                    end_dt = current_shift_date.replace(hour=e_time.hour, minute=e_time.minute, second=0, microsecond=0) + timedelta(days=1)
                                else:
                                    end_dt = current_shift_date.replace(hour=e_time.hour, minute=e_time.minute, second=0, microsecond=0)
                                    
                                start_dt = current_shift_date.replace(hour=s_time.hour, minute=s_time.minute, second=0, microsecond=0)


                                # Store shift data
                                row_shifts.append({
                                    "start_time": start_dt,
                                    "end_time": end_dt,
                                    "location": shift_location
                                })

                            except Exception as e:
                                with open("ocr_debug.log", "a") as f:
                                    f.write(f"DEBUG: Time parse ERROR for '{t_start_str}' / '{t_end_str}': {e}\n")
                                print(f"Time parse error: {e}")
                                # Fallback on error too
                                s_dt = current_shift_date.replace(hour=9, minute=0, second=0, microsecond=0)
                                e_dt = current_shift_date.replace(hour=17, minute=0, second=0, microsecond=0)
                                row_shifts.append({
                                    "start_time": s_dt,
                                    "end_time": e_dt,
                                    "location": shift_location,
                                    "notes": f"RAW: {clean_time_text} (No match)"
                                })
                                continue
                        except Exception as loop_e:
                            print(f"CRITICAL ERROR processing column {col_idx}: {loop_e}")
                            traceback.print_exc()
                            continue

                    row_shifts_count = len(row_shifts)
                    # Flush to make these shifts visible for duplicate checks in subsequent rows
                    session.flush()
                    
                    with open("ocr_debug.log", "a") as f:
                        f.write(f"DEBUG: Loop finished for row '{full_name}'. Row Shifts: {row_shifts_count}\n")

                    print(f"DEBUG: Checking employee: {employee}")
                    if employee:
                        try:
                            # Save shifts
                            print(f"DEBUG: Saving {len(row_shifts)} shifts for {employee.first_name}")
                            for s_data in row_shifts:
                                # Check for existing shift (duplicate prevention)
                                existing_shift = session.exec(select(Shift).where(
                                    Shift.employee_id == employee.id,
                                    Shift.start_time >= s_data["start_time"].replace(hour=0, minute=0, second=0),
                                    Shift.start_time < s_data["start_time"].replace(hour=0, minute=0, second=0) + timedelta(days=1)
                                )).first()
                                
                                if existing_shift:
                                    # Skip duplicate - don't auto-update to preserve manual edits
                                    print(f"DEBUG: Skipped duplicate shift for {employee.first_name} on {s_data['start_time'].date()}")
                                    with open("ocr_debug.log", "a") as f:
                                        f.write(f"DUPLICATE SKIPPED (location preserved): {employee.first_name} - {s_data['start_time']} to {s_data['end_time']}\n")
                                    continue
                                
                                # Create new shift
                                shift = Shift(
                                    employee_id=employee.id,
                                    role_id=employee.default_role_id,
                                    start_time=s_data["start_time"],
                                    end_time=s_data["end_time"],
                                    location=s_data["location"],
                                    notes=s_data.get("notes")
                                )
                                if dry_run:
                                    parsed_shifts.append({
                                        "employee_id": employee.id,
                                        "employee_name": f"{employee.first_name} {employee.last_name}",
                                        "role_id": employee.default_role_id,
                                        "start_time": s_data['start_time'].isoformat(),
                                        "end_time": s_data['end_time'].isoformat(),
                                        "notes": s_data.get("notes", "OCR Import"),
                                        "location": s_data['location'],
                                        "is_vacation": False
                                    })
                                else:
                                    session.add(shift)
                                imported_count += 1
                                
                                # Log for debug
                                with open("ocr_debug.log", "a") as f:
                                    f.write(f"DEBUG: Created Shift for {employee.first_name}: {s_data['start_time']} - {s_data['end_time']} ({s_data['location']})\n")
                        except Exception as save_e:
                            print(f"Error saving shifts for {employee.first_name}: {save_e}")
                            traceback.print_exc()
                            with open("ocr_debug.log", "a") as f:
                                f.write(f"CRITICAL ERROR saving shifts for {employee.first_name}: {save_e}\n")
                    else:
                        # Employee not found - store name AND shifts
                        unmatched_employees.append({
                            "name": full_name,
                            "shifts": [
                                {
                                    "start_time": s['start_time'].isoformat(),
                                    "end_time": s['end_time'].isoformat(),
                                    "location": s['location']
                                } for s in row_shifts
                            ]
                        })
                        
                else:
                    unmatched_lines.append(full_line_text)

            # --- Memory Cleanup per Page ---
            del img
            del img_np
            if 'rotated_img' in locals(): del rotated_img
            if 'img_np_rot' in locals(): del img_np_rot
            if 'curr_results' in locals(): del curr_results
            if 'results' in locals(): del results
            gc.collect()


        # MOVED OUTSIDE LOOP: Commit all shifts after processing all pages
        print(f"DEBUG: Reached end of image loop. Total imported_count: {imported_count}")
        with open("ocr_debug.log", "a") as f:
            f.write(f"DEBUG: Reached end of image loop. Total imported_count: {imported_count}\\n")    
        if not dry_run:
            try:
                # Check session state
                print(f"DEBUG: Session has {len(session.new)} new objects before commit")
                print(f"DEBUG: Attempting to commit {imported_count} shifts to database...")
                with open("ocr_debug.log", "a") as f:
                    f.write(f"DEBUG: Session has {len(session.new)} new objects\\n")
                    f.write(f"DEBUG: Attempting to commit {imported_count} shifts to database...\\n")
                session.commit()
                print(f"DEBUG: Successfully committed {imported_count} shifts!")
                with open("ocr_debug.log", "a") as f:
                    f.write(f"DEBUG: Successfully committed {imported_count} shifts!\\n")
            except Exception as e:
                print(f"DEBUG: COMMIT FAILED: {e}")
                import traceback
                traceback.print_exc()
                with open("ocr_debug.log", "a") as f:
                    f.write(f"DEBUG: COMMIT FAILED: {e}\\n")
                return {"message": "Database error", "errors": [str(e)]}
    
        return {
            "message": f"OCR Processing Complete. Found {imported_count} shifts.",
            "errors": errors,
            "raw_text_preview": extracted_text[:500] + "...", # This will be empty if direct PDF text failed
            "parsed_shifts": parsed_shifts,
            "unmatched_lines": unmatched_lines,
            "unmatched_employees": unmatched_employees  # Already deduplicated by only adding once per name
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"message": "OCR Failed", "errors": [str(e)]}

@app.post("/shifts/bulk/")
def create_shifts_bulk(shifts: List[dict], session: Session = Depends(get_session)):
    count = 0
    for s_data in shifts:
        try:
            # Parse dates
            start = datetime.fromisoformat(s_data['start_time'])
            end = datetime.fromisoformat(s_data['end_time'])
            
            shift = Shift(
                employee_id=s_data['employee_id'],
                role_id=s_data['role_id'],
                start_time=start,
                end_time=end,
                notes=s_data.get('notes'),
                location=s_data.get('location'), # Added location
                is_vacation=s_data.get('is_vacation', False)
            )
            session.add(shift)
            count += 1
        except Exception as e:
            print(f"Error creating shift: {e}")
            continue
            
    try:
        session.commit()
    except Exception as e:
        session.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Database Commit Failed: {str(e)}")
        
    print(f"DEBUG: Bulk create processed {count} shifts.")
    return {"message": f"Created {count} shifts"}

@app.post("/rotations/")
def update_rotation(state: RotationState, session: Session = Depends(get_session)):
    try:
        existing = session.get(RotationState, state.context_key)
        if existing:
            existing.last_employee_id = state.last_employee_id
            existing.updated_at = datetime.utcnow()
            session.add(existing)
        else:
            session.add(state)
        session.commit()
        return {"status": "success", "message": f"Rotation updated for {state.context_key}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/shifts/{shift_id}/call-sheet")
def get_call_sheet(shift_id: int, session: Session = Depends(get_session)):
    try:
        # ... (Existing logic for Shift, Dates) ...
        target_shift = session.get(Shift, shift_id)
        if not target_shift: raise HTTPException(status_code=404, detail="Shift not found")
        
        # Target Week (Saturday to Saturday)
        start_dt = target_shift.start_time
        # Saturday is weekday 5. If today is Mon (0), we go back 2 days. If Sat (5), stay. If Sun (6), go back 1.
        days_since_saturday = (start_dt.weekday() + 2) % 7  # Sat=0, Sun=1, Mon=2, ..., Fri=6
        start_of_week = start_dt.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=days_since_saturday)
        end_of_week = start_of_week + timedelta(days=7)
        
        # Target Day info
        day_start = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        # Candidates
        cand_role_id = target_shift.role_id
    
        # Whitelist
        ALLOWED_ROLES = [3, 4, 7, 8]
        if cand_role_id not in ALLOWED_ROLES:
            raise HTTPException(status_code=400, detail="Call Sheet not available for this role type.")
        
        candidate_objects = []
        
        # Helper: Check if employee has Supervisor as secondary role
        def has_supervisor_role(emp):
            """Returns True if employee has Supervisor (role_id=5) as a secondary role."""
            return any(r.id == 5 for r in emp.roles)
        
        if cand_role_id == 4:
            # Maintenance
            pt_maint = [e for e in session.exec(select(Employee).where(Employee.default_role_id == 4, Employee.is_full_time == False)).all() if not has_supervisor_role(e)]
            ft_maint = [e for e in session.exec(select(Employee).where(Employee.default_role_id == 4, Employee.is_full_time == True)).all() if not has_supervisor_role(e)]
            
            pt_maint.sort(key=lambda x: x.hire_date or datetime.max)
            ft_maint.sort(key=lambda x: x.hire_date or datetime.max)
            
            # Apply Rotation for FT Maintenance
            rot_state = session.get(RotationState, "maint_ft")
            if rot_state and ft_maint:
                # Rotate list
                try:
                    # Find index of last called
                    idx = next(i for i, emp in enumerate(ft_maint) if emp.id == rot_state.last_employee_id)
                    # Rotate: Start from idx + 1
                    ft_maint = ft_maint[idx+1:] + ft_maint[:idx+1]
                except StopIteration:
                    pass # Last called person not found in list (maybe deleted/changed role), keep default order
            
            # Pre-fetch shifts for OT calculation
            all_maint = pt_maint + ft_maint
            maint_ids = [m.id for m in all_maint]
            maint_week_shifts = session.exec(select(Shift).where(Shift.employee_id.in_(maint_ids), Shift.start_time >= start_of_week, Shift.start_time < end_of_week)).all()
            maint_shifts_map = {mid: [] for mid in maint_ids}
            for s in maint_week_shifts: 
                maint_shifts_map[s.employee_id].append(s)
            
            target_duration = (target_shift.end_time - target_shift.start_time).total_seconds() / 3600
            # Apply 30min lunch deduction for target if >= 7.5h
            effective_target = target_duration - 0.5 if target_duration >= 7.5 else target_duration
            
            # Split PT Maint into standard vs OT
            pt_standard = []
            pt_ot = []
            for emp in pt_maint:
                shifts = maint_shifts_map.get(emp.id, [])
                weekly_hours = 0
                for s in shifts:
                    dur = (s.end_time - s.start_time).total_seconds() / 3600
                    if dur >= 7.5: dur -= 0.5  # Lunch deduction
                    weekly_hours += dur
                
                if weekly_hours + effective_target > 40:
                    pt_ot.append((emp, "Part Time Maintenance (OT)"))
                else:
                    pt_standard.append((emp, "Part Time Maintenance"))
            
            # Split FT Maint into standard vs OT
            ft_standard = []
            ft_ot = []
            for emp in ft_maint:
                shifts = maint_shifts_map.get(emp.id, [])
                weekly_hours = 0
                for s in shifts:
                    dur = (s.end_time - s.start_time).total_seconds() / 3600
                    if dur >= 7.5: dur -= 0.5  # Lunch deduction
                    weekly_hours += dur
                
                if weekly_hours + effective_target > 40:
                    ft_ot.append((emp, "Full Time Maintenance (OT)"))
                else:
                    ft_standard.append((emp, "Full Time / Probationary Maintenance"))
            
            # Order: PT standard, FT standard, PT OT, FT OT
            candidate_objects = pt_standard + ft_standard + pt_ot + ft_ot
        else:
            # Cashier - filter out employees with Supervisor as secondary role and inactive employees
            all_cashiers = [e for e in session.exec(select(Employee).where(Employee.default_role_id.in_([3, 7, 8]))).all() if not has_supervisor_role(e) and e.is_active != False]
            
            # Helper function for consistent FT check
            def is_full_time(emp):
                return emp.is_full_time == True or emp.is_full_time == 1
            
            # Pre-fetch shifts
            cashier_ids = [c.id for c in all_cashiers]
            all_week_shifts = session.exec(select(Shift).where(Shift.employee_id.in_(cashier_ids), Shift.start_time >= start_of_week, Shift.start_time < end_of_week)).all()
            shifts_map = {cid: [] for cid in cashier_ids}
            for s in all_week_shifts: shifts_map[s.employee_id].append(s)
            
            target_duration = (target_shift.end_time - target_shift.start_time).total_seconds() / 3600
            
            # Calculate weekly hours for each employee
            def get_weekly_hours(emp_id):
                shifts = shifts_map.get(emp_id, [])
                return sum([(s.end_time - s.start_time).total_seconds()/3600 for s in shifts])
            
            # Build all 3 pages
            page_1_list = []  # PT + FT under 40h
            page_2_list = []  # FT at/over 40h (OT)
            page_3_list = []  # ALL PT for OT (no FT)
            
            # Sort by hire date
            sorted_cashiers = sorted(all_cashiers, key=lambda x: x.hire_date or datetime.max)
            
            for emp in sorted_cashiers:
                weekly_hours = round(get_weekly_hours(emp.id), 1)  # Round to fix floating point precision
                is_ft = is_full_time(emp)
                
                # FT with <= 32h scheduled is treated as PT for call sheet
                treat_as_pt = (not is_ft) or (is_ft and weekly_hours <= 32)
                
                if treat_as_pt:
                    # PT employee (or FT with <= 32h) -> Page 1 AND Page 3
                    page_1_list.append((emp, "Page 1"))
                    page_3_list.append((emp, "Page 3"))
                elif is_ft:
                    # FT employee with > 32h
                    if weekly_hours >= 40 or (weekly_hours + target_duration) > 40:
                        # FT at or over 40h -> Page 2 (OT)
                        page_2_list.append((emp, "Page 2"))
                    else:
                        # FT between 32-40h -> Page 1 only
                        page_1_list.append((emp, "Page 1"))

            candidate_objects = page_1_list + page_2_list + page_3_list
        
        if not candidate_objects:
            return []
        
        # Calculate Final Results
        cand_ids = [c[0].id for c in candidate_objects]
        week_shifts = session.exec(select(Shift).where(Shift.employee_id.in_(cand_ids), Shift.start_time >= start_of_week, Shift.start_time < end_of_week)).all()
        
        emp_shifts = {cid: [] for cid in cand_ids}
        for s in week_shifts:
            if s.employee_id: emp_shifts[s.employee_id].append(s)
            
        # Target duration logic (Maintenance Unpaid Meal)
        raw_target_duration = (target_shift.end_time - target_shift.start_time).total_seconds() / 3600
        effective_target_duration = raw_target_duration
        
        target_shift_notes = []
        if target_shift.role_id == 4 and raw_target_duration >= 7.5:
             effective_target_duration -= 0.5
             target_shift_notes.append("30m Unpaid Lunch")
        
        results = []
        rank = 1
        
        for emp, section in candidate_objects:
            shifts = emp_shifts.get(emp.id, [])
            
            # Recalculate Weekly Hours with Deduction Logic for existing Maintenance Shifts
            weekly_hours = 0
            for s in shifts:
                 duration = (s.end_time - s.start_time).total_seconds() / 3600
                 # If it's a Maintenance shift (Role 4) >= 7.5h, deduct 0.5h
                 if s.role_id == 4 and duration >= 7.5:
                     duration -= 0.5
                 weekly_hours += duration
            
            overlap_duration = 0
            working_today = False
            daily_hours = 0
            
            for s in shifts:
                if s.start_time < target_shift.end_time and s.end_time > target_shift.start_time:
                    latest_start = max(s.start_time, target_shift.start_time)
                    earliest_end = min(s.end_time, target_shift.end_time)
                    delta = (earliest_end - latest_start).total_seconds() / 3600
                    if delta > 0: overlap_duration += delta
                
                if s.start_time >= day_start and s.end_time < day_end:
                    working_today = True
                    # Recalculate Daily Hours with Deduction Logic
                    duration = (s.end_time - s.start_time).total_seconds() / 3600
                    if s.role_id == 4 and duration >= 7.5:
                         duration -= 0.5
                    daily_hours += duration
            
            status = "Available"
            details = ""
            
            # Check if this is the employee who originally had the shift (Called Out)
            if target_shift.employee_id and emp.id == target_shift.employee_id:
                status = "CO"
                details = "Called Out - Original shift holder"
            elif overlap_duration > 0:
                status = "Working"
                details = f"Overlap {overlap_duration:.1f}h"
            elif working_today:
                 status = "Working"
                 details = f"Shift today ({daily_hours:.1f}h)"
                 
                 # Daily Limit Logic
                 if section == "Part Time Cashiers (Priority)":
                     limit = 8
                     if daily_hours + effective_target_duration > limit:
                         status = "OT"
                         details = f"Daily > {limit}h ({daily_hours+effective_target_duration:.1f}h)"
            
            # Global 16h Safety Warning
            if daily_hours + effective_target_duration > 16:
                details += f" Warning: >16h ({daily_hours+effective_target_duration:.1f}h)"

            # Weekly OT Logic
            if weekly_hours + effective_target_duration > 40:
                status = "OT"
                details = f"Weekly > 40h ({weekly_hours+effective_target_duration:.1f}h)"
            
            # Append notes about meal break to details if applicable
            if target_shift_notes:
                 details += " (" + ", ".join(target_shift_notes) + ")"

            # Note-Based Constraints Logic
            violation_reason = None  # Initialize before checking
            
            # Helper to determine shift type based on hour
            # 3rd: midnight-6am, 1st: 6am-2pm, 2nd: 2pm-midnight
            def get_shift_type(hour, role_id=None):
                if 0 <= hour < 6:
                    return '3rd'
                elif 6 <= hour < 14:
                    return '1st'
                else:
                    return '2nd'
            
            # Check the structured no_overtime field first
            if emp.no_overtime:
                violation_reason = "Restricted: No Overtime"
            
            # Check the structured no_plaza field
            if not violation_reason and emp.no_plaza:
                if target_shift.location == "Plaza":
                    violation_reason = "Restricted: No Plaza"
            
            # Check availability_grid (JSON: {day: {shift: bool}})
            if not violation_reason and emp.availability_grid:
                import json
                try:
                    grid = json.loads(emp.availability_grid)
                    # Determine day of week (sun, mon, tue, etc.)
                    day_names = ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun']
                    shift_day = day_names[target_shift.start_time.weekday()]
                    
                    # Determine shift type
                    hour = target_shift.start_time.hour
                    shift_type = get_shift_type(hour, target_shift.role_id)
                    
                    # Check if available
                    if shift_day in grid and shift_type in grid[shift_day]:
                        if not grid[shift_day][shift_type]:
                            violation_reason = f"Unavailable: {shift_day.upper()} {shift_type} Shift"
                except (json.JSONDecodeError, KeyError):
                    pass  # Invalid JSON, ignore
            
            # Legacy Note-Based Constraints Logic (fallback)
            if not violation_reason and emp.notes:
                notes_lower = emp.notes.lower()
                notes_upper = emp.notes.upper()
                
                # Get target shift info
                shift_hour = target_shift.start_time.hour
                shift_day = target_shift.start_time.strftime('%a').upper()[:3]  # SUN, MON, TUE, etc.
                
                # Determine which shift type this is (3rd: midnight-6am, 1st: 6am-2pm, 2nd: 2pm-midnight)
                if 0 <= shift_hour < 6:
                    shift_type = "3RD"
                elif 6 <= shift_hour < 14:
                    shift_type = "1ST"
                else:
                    shift_type = "2ND"
                
                # Parse days from notes (e.g., "AVAIL SUN,MON", "TUE-SAT", "TUE THUR FRI SAT")
                day_names = ['SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
                day_aliases = {'THUR': 'THU', 'THURS': 'THU', 'SUNDAY': 'SUN', 'MONDAY': 'MON', 'TUESDAY': 'TUE', 'WEDNESDAY': 'WED', 'THURSDAY': 'THU', 'FRIDAY': 'FRI', 'SATURDAY': 'SAT'}
                
                available_days = set()
                available_shifts = set()
                has_day_restriction = False
                has_shift_restriction = False
                
                # Check for "NO PLAZA"
                if "no plaza" in notes_lower and target_shift.location == "Plaza":
                    violation_reason = "Restricted: No Plaza"
                
                # Check for "NO OVERTIME"
                elif "no overtime" in notes_lower or "no ot" in notes_lower or "do not call for overtime" in notes_lower or "do not call for extra" in notes_lower:
                    violation_reason = "Restricted: No Overtime"
                
                # Check for "1ST SHIFT ONLY", "2ND SHIFT ONLY", "3RD SHIFT ONLY"
                elif "1st shift only" in notes_lower or "1st only" in notes_lower:
                    has_shift_restriction = True
                    available_shifts.add("1ST")
                elif "2nd shift only" in notes_lower or "2nd only" in notes_lower:
                    has_shift_restriction = True
                    available_shifts.add("2ND")
                elif "3rd shift only" in notes_lower or "3rd only" in notes_lower:
                    has_shift_restriction = True
                    available_shifts.add("3RD")
                
                # Check for "AVAIL 2ND & 3RD", "AVAIL 1ST AND 2ND", etc.
                elif "avail" in notes_lower:
                    # Parse shifts mentioned
                    if "1st" in notes_lower:
                        available_shifts.add("1ST")
                        has_shift_restriction = True
                    if "2nd" in notes_lower:
                        available_shifts.add("2ND")
                        has_shift_restriction = True
                    if "3rd" in notes_lower:
                        available_shifts.add("3RD")
                        has_shift_restriction = True
                    
                    # Parse days mentioned
                    for day in day_names:
                        if day in notes_upper:
                            available_days.add(day)
                            has_day_restriction = True
                    for alias, day in day_aliases.items():
                        if alias in notes_upper:
                            available_days.add(day)
                            has_day_restriction = True
                
                # Check for "AFTER 4PM", "AFTER 5PM" etc.
                import re
                after_match = re.search(r'after\s*(\d+)\s*(pm|am)?', notes_lower)
                if after_match:
                    after_hour = int(after_match.group(1))
                    is_pm = after_match.group(2) == 'pm' if after_match.group(2) else (after_hour < 12)
                    if is_pm and after_hour < 12:
                        after_hour += 12
                    if shift_hour < after_hour:
                        violation_reason = f"Unavailable: Only after {after_match.group(1)}{after_match.group(2) or 'PM'}"
                
                # Apply day restriction
                if has_day_restriction and not violation_reason:
                    if shift_day not in available_days:
                        violation_reason = f"Unavailable: Not avail {shift_day}"
                
                # Apply shift restriction
                if has_shift_restriction and not violation_reason:
                    if shift_type not in available_shifts:
                        violation_reason = f"Unavailable: Not avail {shift_type} Shift"
                
                if violation_reason:
                    status = "Unavailable"
                    # Prepend restriction to details for visibility
                    details = f"{violation_reason}. " + details
            
            # Determine Answer Field Value
            answer_val = ""
            if violation_reason:
                answer_val = "No"
            elif overlap_duration > 0:
                answer_val = f"OL {overlap_duration:.1f}h"
            elif section.startswith("Page 2") or section.startswith("Page 3"):
                # OT pages (Page 2 & 3) - don't mark as OT since they're on an OT page
                if status == "Working":
                    answer_val = "W"
                # else leave blank - they're expected to work OT
            elif weekly_hours + effective_target_duration > 40:
                answer_val = "Over 40"
            elif status == "Working":
                answer_val = "W"

            entry = {
                "rank": rank,
                "section": section,
                "id": emp.id,
                "name": f"{emp.first_name} {emp.last_name}",
                "phone": emp.phone,
                "hire_date": emp.hire_date.isoformat() if emp.hire_date else None,
                "notes": emp.notes,
                "status": status,
                "details": details.strip(),
                "weekly_hours": round(weekly_hours, 1),
                "answer": answer_val
            }
            results.append(entry)
            rank += 1
            
        return results
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
